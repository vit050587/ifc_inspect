#!/usr/bin/env python3
"""
Скрипт парсинга IFC файлов.
Создает сводную таблицу по типам элементов, материалам с количеством и объемом.
Для несущих элементов из монолитного железобетона (стены, перекрытия, колонны, фундаменты)
извлекает характеристики материала: класс бетона, морозостойкость, водонепроницаемость.

Использует прямое чтение IFC через ifcopenshell (логика из extract.py).

Формат выходной таблицы:
Тип (RU) | Тип элемента | Материал (с характеристиками: Бетон В30 F150 W6) | Количество, шт | Объем, м³
"""

import os
import re
import json
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple
from collections import defaultdict
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import ifcopenshell
import ifcopenshell.util.element
import ifcopenshell.util.unit


# Несущие элементы из монолитного железобетона
LOAD_BEARING_CONCRETE_ELEMENTS = {
    'IfcWall', 'IfcColumn', 'IfcSlab', 'IfcFooting', 'IfcPile', 
    'IfcBeam', 'IfcMember', 'IfcPlate', 'IfcStair', 'IfcStairFlight',
    'IfcRamp'
}

# Mapping IFC классов к русским названиям типов элементов
IFC_TYPE_RU_MAPPING = {
    'IfcWall': 'Стены',
    'IfcColumn': 'Колонны',
    'IfcSlab': 'Перекрытия',
    'IfcStair': 'Лестницы',
    'IfcStairFlight': 'Лестничные_марши',
    'IfcRamp': 'Пандусы',
    'IfcBeam': 'Балки',
    'IfcFooting': 'Фундаменты',
    'IfcPile': 'Сваи',
    'IfcRoof': 'Крыши',
    'IfcPlate': 'Плиты',
    'IfcMember': 'Элементы_каркаса',
    'IfcCurtainWall': 'Навесные_стены',
    'IfcWindow': 'Окна',
    'IfcDoor': 'Двери',
    'IfcFurnishingElement': 'Мебель',
    'IfcBuildingElementProxy': 'Прочие_элементы',
    'IfcOpeningElement': 'Проемы',
    'IfcReinforcingMesh': 'Арматурные_сетки',
    'IfcBuildingStorey': 'Этажи',
    'IfcBuilding': 'Здание',
    'IfcRailing': 'Ограждения',
    'IfcCovering': 'Покрытия',
}


def get_unit_scale(ifc_file):
    """Получить масштаб единиц из IFC файла"""
    try:
        units = ifcopenshell.util.unit.get_unit_scale(ifc_file)
        return units
    except:
        return 1.0


def calculate_volume_by_geometry(shape, unit_scale):
    """Расчет объема через bounding box с учетом масштаба"""
    try:
        bbox_min = shape.bounding_box.lower_corner
        bbox_max = shape.bounding_box.upper_corner
        
        dx = (bbox_max[0] - bbox_min[0]) * unit_scale
        dy = (bbox_max[1] - bbox_min[1]) * unit_scale
        dz = (bbox_max[2] - bbox_min[2]) * unit_scale
        
        return abs(dx * dy * dz)
    except Exception as e:
        return 0.0


def calculate_area_by_geometry(shape, unit_scale):
    """Расчет площади поверхности через геометрию"""
    try:
        area = shape.geometry.surface_area
        return area * (unit_scale ** 2)
    except Exception as e:
        return 0.0


def get_property_value(psets, pset_name, prop_name):
    """Получить значение свойства из psets"""
    if pset_name in psets:
        props = psets[pset_name]
        if prop_name in props:
            val = props[prop_name]
            if val is not None and prop_name != 'id':
                return val
    return None


def find_ifc_report_file(session_folder: str) -> Optional[str]:
    """Находит файл ifc_report.xlsx в папке сессии."""
    report_path = Path(session_folder) / 'ifc_report.xlsx'
    
    if not report_path.exists():
        print(f"⚠️  Файл ifc_report.xlsx не найден: {report_path}")
        return None
    
    return str(report_path)


def extract_concrete_properties_from_name(material_name: str) -> Tuple[str, Optional[str], Optional[str], Optional[str]]:
    """
    Извлекает характеристики бетона из названия материала.
    
    Примеры входных данных:
    - "Бетон В30 F150 W8"
    - "Бетон B35 W6 F150 Гранит"
    - "Бетон В7.5 неармированный"
    - "02_Бетон B30 W6 F150 Гравий"
    
    Returns:
        Tuple[str, Optional[str], Optional[str], Optional[str]]: 
            (очищенное_имя_материала, класс_бетона, морозостойкость, водонепроницаемость)
    """
    if not material_name or not isinstance(material_name, str):
        return material_name, None, None, None
    
    original_name = material_name.strip()
    
    # Паттерны для извлечения характеристик
    # Класс бетона: В20, В25, В30, В35, В40, B10, B15, B20, B25, B30, B35, B40 (русская или латинская В)
    grade_pattern = r'\b([BВ]\d+(?:\.\d+)?)\b'
    # Морозостойкость: F100, F150, F200, F250, F300
    freeze_pattern = r'\b(F\d+)\b'
    # Водонепроницаемость: W4, W6, W8, W10, W12, W14, W20
    water_pattern = r'\b(W\d+)\b'
    
    concrete_grade = None
    freeze_durability = None
    water_resist = None
    
    # Ищем характеристики в названии
    grade_match = re.search(grade_pattern, original_name)
    if grade_match:
        concrete_grade = grade_match.group(1)
        # Нормализуем: заменяем латинскую B на русскую В
        if concrete_grade.startswith('B') and not concrete_grade.startswith('В'):
            concrete_grade = 'В' + concrete_grade[1:]
    
    freeze_match = re.search(freeze_pattern, original_name)
    if freeze_match:
        freeze_durability = freeze_match.group(1)
    
    water_match = re.search(water_pattern, original_name)
    if water_match:
        water_resist = water_match.group(1)
    
    # Очищаем название от характеристик для базового имени
    clean_name = original_name
    # Удаляем найденные характеристики из названия
    if grade_match:
        clean_name = re.sub(re.escape(grade_match.group(0)), '', clean_name)
    if freeze_match:
        clean_name = re.sub(re.escape(freeze_match.group(0)), '', clean_name)
    if water_match:
        clean_name = re.sub(re.escape(water_match.group(0)), '', clean_name)
    
    # Удаляем лишние пробелы и спецсимволы
    clean_name = re.sub(r'\s+', ' ', clean_name).strip()
    clean_name = re.sub(r'^02_|^03_|^30_|^31_', '', clean_name)  # Удаляем префиксы номеров групп
    clean_name = clean_name.strip('_').strip()
    
    # Если материал содержит "Бетон" и есть характеристики - формируем полное имя
    if 'бетон' in clean_name.lower() or 'раствор' in clean_name.lower():
        if concrete_grade or freeze_durability or water_resist:
            # Формируем итоговое название: "Бетон В30 F150 W6"
            parts = [clean_name] if clean_name else ['Бетон']
            if concrete_grade:
                parts.append(concrete_grade)
            if freeze_durability:
                parts.append(freeze_durability)
            if water_resist:
                parts.append(water_resist)
            return ' '.join(parts), concrete_grade, freeze_durability, water_resist
    
    return clean_name if clean_name else original_name, concrete_grade, freeze_durability, water_resist


def get_material_with_properties(row: pd.Series, ifc_class: str, material_col: str) -> Tuple[str, Optional[str], Optional[str], Optional[str]]:
    """
    Получает полное имя материала с характеристиками бетона.
    
    Логика:
    1. Берем материал из указанной колонки
    2. Пытаемся извлечь характеристики из названия материала (regex)
    3. Если характеристик нет в названии, ищем в стандартных колонках MGE_*
    4. Возвращаем комбинированное имя типа "Бетон В30 F150 W6"
    
    Returns:
        Tuple[str, Optional[str], Optional[str], Optional[str]]: 
            (материал_с_характеристиками, класс_бетона, морозостойкость, водонепроницаемость)
    """
    # Получаем базовый материал
    base_material = 'Не указан'
    if material_col in row.index:
        mat_val = row.get(material_col)
        if mat_val and str(mat_val) != '0' and str(mat_val) != 'nan' and str(mat_val).strip():
            base_material = str(mat_val).strip()
    
    # Если материал не указан, пытаемся найти альтернативную колонку MGE_Material
    if base_material == 'Не указан' or base_material == '0':
        for col in row.index:
            if 'MGE_Material' in col and ifc_class[3:] in col:
                mat_val = row.get(col)
                if mat_val and str(mat_val) != '0' and str(mat_val) != 'nan' and str(mat_val).strip():
                    base_material = str(mat_val).strip()
                    break
    
    # Сначала пробуем извлечь характеристики из названия материала
    clean_material, grade_from_name, freeze_from_name, water_from_name = \
        extract_concrete_properties_from_name(base_material)
    
    # Если характеристики найдены в названии - возвращаем их
    if grade_from_name or freeze_from_name or water_from_name:
        return clean_material, grade_from_name, freeze_from_name, water_from_name
    
    # Если характеристик нет в названии, ищем в стандартных колонках MGE_*
    concrete_grade = None
    freeze_durability = None
    water_resist = None
    
    # Ищем морозостойкость F***
    freeze_col = 'ExpCheck_MaterialConcrete:MGE_FreezeDurability'
    if freeze_col in row.index:
        freeze_val = row.get(freeze_col)
        if freeze_val and str(freeze_val) != '0' and str(freeze_val) != 'nan' and str(freeze_val).strip():
            freeze_durability = str(freeze_val).strip()
    
    # Ищем водонепроницаемость W**
    water_col = 'ExpCheck_MaterialConcrete:MGE_WaterResist'
    if water_col in row.index:
        water_val = row.get(water_col)
        if water_val and str(water_val) != '0' and str(water_val) != 'nan' and str(water_val).strip():
            water_resist = str(water_val).strip()
    
    # Ищем класс бетона
    grade_col = 'ExpCheck_MaterialConcrete:MGE_ConcreteGrade'
    if grade_col in row.index:
        grade_val = row.get(grade_col)
        if grade_val and str(grade_val) != '0' and str(grade_val) != 'nan' and str(grade_val).strip():
            concrete_grade = str(grade_val).strip()
    
    # Если марки нет в стандартной колонке, пытаемся извлечь из Element Specific:Name
    if not concrete_grade:
        name_col = 'Element Specific:Name'
        if name_col in row.index:
            elem_name = row.get(name_col)
            if elem_name and str(elem_name) != 'nan' and str(elem_name).strip():
                elem_name = str(elem_name)
                grade_match = re.search(r'\b([BВ]\d+(?:\.\d+)?)\b', elem_name)
                if grade_match:
                    concrete_grade = grade_match.group(1)
                    if concrete_grade.startswith('B') and not concrete_grade.startswith('В'):
                        concrete_grade = 'В' + concrete_grade[1:]
    
    # Если все еще нет марки, пробуем ObjectType
    if not concrete_grade:
        type_col = 'Element Specific:ObjectType'
        if type_col in row.index:
            obj_type = row.get(type_col)
            if obj_type and str(obj_type) != 'nan' and str(obj_type).strip():
                obj_type = str(obj_type)
                grade_match = re.search(r'\b([BВ]\d+(?:\.\d+)?)\b', obj_type)
                if grade_match:
                    concrete_grade = grade_match.group(1)
                    if concrete_grade.startswith('B') and not concrete_grade.startswith('В'):
                        concrete_grade = 'В' + concrete_grade[1:]
    
    # Формируем итоговое название материала с характеристиками
    final_material = clean_material
    if any(x in clean_material.lower() for x in ['бетон', 'раствор']):
        parts = [clean_material] if clean_material else ['Бетон']
        if concrete_grade and concrete_grade not in clean_material:
            parts.append(concrete_grade)
        if freeze_durability and freeze_durability not in clean_material:
            parts.append(freeze_durability)
        if water_resist and water_resist not in clean_material:
            parts.append(water_resist)
        final_material = ' '.join(parts)
    
    return final_material, concrete_grade, freeze_durability, water_resist


def parse_excel_report(excel_path: str) -> Dict[str, Any]:
    """
    Парсит Excel файл отчета IFC (лист "Элементы").
    Извлекает данные для сводной таблицы: Тип, IfcClass, Материал, Количество, Объем, Площадь.
    
    Логика:
    1. Для каждой строки определяем IfcClass и LongName
    2. Ищем колонки с материалами по паттерну Qto_*:NN_Материал
    3. Для каждого материала берем значения (это уже объем/площадь в зависимости от типа)
    4. Также берем NetVolume, NetSideArea, OuterSurfaceArea и т.д. как дополнительные метрики
    5. Агрегируем по (Тип RU, IfcClass, Материал) с учетом марки бетона
    """
    
    print(f"\n📊 Парсинг Excel отчета IFC: {os.path.basename(excel_path)}")
    
    try:
        # Читаем лист "Элементы"
        df = pd.read_excel(excel_path, sheet_name='Элементы', dtype=str)
        
        # Нормализуем имена колонок
        df.columns = [str(col).strip() for col in df.columns]
        
        # Ожидаемые имена колонок
        col_long_name = 'Element Specific:LongName'
        col_ifc_class = 'Ifc Class'
        
        # Проверяем наличие обязательных колонок
        if col_ifc_class not in df.columns:
            raise ValueError(f"Столбец '{col_ifc_class}' не найден в файле.")
        if col_long_name not in df.columns:
            col_long_name = 'Element Specific:Name'
            if col_long_name not in df.columns:
                raise ValueError(f"Столбец '{col_long_name}' не найден в файле")
        
        # Находим все колонки с материалами по паттерну Qto_*:цифра_название
        material_pattern = re.compile(r'^(Qto_\w+):(\d+)_(.+)$')
        material_columns = []
        
        for col in df.columns:
            match = material_pattern.match(col)
            if match:
                qto_type = match.group(1)  # например Qto_WallBaseQuantities
                num = match.group(2)        # номер группы материалов
                material_name = match.group(3).strip()
                material_columns.append({
                    'column': col,
                    'qto_type': qto_type,
                    'number': num,
                    'material_name': material_name
                })
        
        print(f"   📋 Найдено {len(material_columns)} колонок с материалами")
        
        # Колонки с объемами и площадями
        volume_cols_pattern = re.compile(r'^Qto_\w+:(NetVolume|GrossVolume)$')
        area_cols_pattern = re.compile(r'^Qto_\w+:(NetSideArea|NetArea|GrossSideArea|OuterSurfaceArea)$')
        
        volume_cols = [c for c in df.columns if volume_cols_pattern.match(c)]
        area_cols = [c for c in df.columns if area_cols_pattern.match(c)]
        
        print(f"   📏 Найдено {len(volume_cols)} колонок объема, {len(area_cols)} колонок площади")
        
        # Структура для агрегации: key = (type_ru, ifc_class, material, concrete_grade, freeze_durability, water_resist)
        aggregated_data = defaultdict(lambda: {
            'count': 0,
            'volume': 0.0,
            'area': 0.0,
        })
        
        processed_rows = 0
        
        for idx, row in df.iterrows():
            ifc_class = row.get(col_ifc_class)
            if not ifc_class or pd.isna(ifc_class):
                continue
            
            ifc_class = str(ifc_class).strip()
            
            # Пропускаем служебные классы
            if ifc_class in ['IfcBuilding', 'IfcBuildingStorey']:
                continue
            
            type_ru = get_type_ru(ifc_class)
            
            # Флаг: была ли найдена хотя бы одна запись для этой строки
            has_material_data = False
            
            # Получаем характеристики бетона из строки (для использования в обоих случаях)
            concrete_grade = None
            freeze_durability = None
            water_resist = None
            
            # Ищем характеристики бетона в стандартных колонках
            grade_col = 'ExpCheck_MaterialConcrete:MGE_ConcreteGrade'
            if grade_col in row.index:
                grade_val = row.get(grade_col)
                if grade_val and str(grade_val) != '0' and str(grade_val) != 'nan' and str(grade_val).strip():
                    concrete_grade = str(grade_val).strip()
            
            freeze_col = 'ExpCheck_MaterialConcrete:MGE_FreezeDurability'
            if freeze_col in row.index:
                freeze_val = row.get(freeze_col)
                if freeze_val and str(freeze_val) != '0' and str(freeze_val) != 'nan' and str(freeze_val).strip():
                    freeze_durability = str(freeze_val).strip()
            
            water_col = 'ExpCheck_MaterialConcrete:MGE_WaterResist'
            if water_col in row.index:
                water_val = row.get(water_col)
                if water_val and str(water_val) != '0' and str(water_val) != 'nan' and str(water_val).strip():
                    water_resist = str(water_val).strip()
            
            # Обрабатываем колонки с материалами
            for mat_col_info in material_columns:
                col_name = mat_col_info['column']
                material_name = mat_col_info['material_name']
                
                if col_name not in df.columns:
                    continue
                
                value = row.get(col_name)
                if value and str(value) != '0' and str(value) != 'nan' and str(value).strip():
                    try:
                        numeric_value = float(str(value).replace(',', '.'))
                        if numeric_value <= 0:
                            continue
                        
                        # Ключ агрегации с характеристиками бетона
                        key = (type_ru, ifc_class, material_name, concrete_grade, freeze_durability, water_resist)
                        
                        # Определяем тип величины по номеру группы материалов
                        # 02 - бетон и подобные (объем), 03 - растворы (объем), 
                        # 30 - гидроизоляция/изоляционные материалы (обычно площадь или объем)
                        # 31 - утеплитель (объем)
                        mat_num = mat_col_info['number']
                        
                        if mat_num in ['02', '03', '31']:  # Объемные материалы
                            aggregated_data[key]['volume'] += numeric_value
                        elif mat_num == '30':  # Изоляционные - может быть и объем и площадь
                            # По умолчанию считаем объемом, если название содержит "мембрана", "изоляция" - это площадь
                            if any(x in material_name.lower() for x in ['мембрана', 'изоляц', 'покрытие', 'штукатурк', 'плитк']):
                                aggregated_data[key]['area'] += numeric_value
                            else:
                                aggregated_data[key]['volume'] += numeric_value
                        else:
                            # По умолчанию - объем
                            aggregated_data[key]['volume'] += numeric_value
                        
                        aggregated_data[key]['count'] += 1
                        has_material_data = True
                        
                    except (ValueError, TypeError):
                        pass
            
            # Если нет данных по материалам, но есть объем/площадь в стандартных колонках
            if not has_material_data:
                # Пытаемся найти материал из ExpCheck_*:MGE_Material
                base_material_col = f'ExpCheck_{ifc_class[3:]}:MGE_Material'
                
                # Используем новую функцию для получения материала с характеристиками
                material_result = get_material_with_properties(row, ifc_class, base_material_col)
                base_material = material_result[0]
                concrete_grade = material_result[1]
                freeze_durability = material_result[2]
                water_resist = material_result[3]
                
                # Получаем объем и площадь из стандартных колонок
                volume = 0.0
                area = 0.0
                
                for vol_col in volume_cols:
                    val = row.get(vol_col)
                    if val and str(val) != '0' and str(val) != 'nan':
                        try:
                            volume += float(str(val).replace(',', '.'))
                        except (ValueError, TypeError):
                            pass
                
                for area_col in area_cols:
                    val = row.get(area_col)
                    if val and str(val) != '0' and str(val) != 'nan':
                        try:
                            area += float(str(val).replace(',', '.'))
                        except (ValueError, TypeError):
                            pass
                
                # Если есть объем или площадь, добавляем запись
                if volume > 0 or area > 0:
                    key = (type_ru, ifc_class, base_material, concrete_grade, freeze_durability, water_resist)
                    aggregated_data[key]['volume'] += volume
                    aggregated_data[key]['area'] += area
                    aggregated_data[key]['count'] += 1
                    has_material_data = True
                
                # Если нет объема/площади, но элемент существует - считаем как штуки
                if not has_material_data:
                    key = (type_ru, ifc_class, base_material, concrete_grade, freeze_durability, water_resist)
                    aggregated_data[key]['count'] += 1
            
            processed_rows += 1
        
        print(f"   ✅ Обработано {processed_rows} строк")
        print(f"   📊 Получено {len(aggregated_data)} уникальных комбинаций (тип, класс, материал)")
        
        # Преобразуем в список для удобства
        items = []
        for (type_ru, ifc_class, material, concrete_grade, freeze_durability, water_resist), data in aggregated_data.items():
            items.append({
                'type_ru': type_ru,
                'ifc_class': ifc_class,
                'material': material,
                'concrete_grade': concrete_grade if concrete_grade else '-',
                'freeze_durability': freeze_durability if freeze_durability else '-',
                'water_resist': water_resist if water_resist else '-',
                'count': int(data['count']),
                'volume': round(data['volume'], 3) if data['volume'] > 0 else '-',
                'area': round(data['area'], 3) if data['area'] > 0 else '-',
            })
        
        # Сортируем: сначала по типу, потом по классу, потом по материалу
        items.sort(key=lambda x: (x['type_ru'], x['ifc_class'], x['material']))
        
        return {
            'success': True,
            'items': items,
            'total_items': len(items),
        }
        
    except Exception as e:
        print(f"   ❌ Ошибка при парсинге Excel: {e}")
        import traceback
        traceback.print_exc()
        return {'success': False, 'error': str(e), 'items': [], 'total_items': 0}


def create_summary_excel(items: List[Dict[str, Any]], output_path: str) -> str:
    """
    Создаёт Excel отчёт со сводной таблицей в формате:
    Тип (RU) | Тип элемента | Материал (с характеристиками: Бетон В30 F150 W6) | Количество, шт | Объем, м³ | Площадь, м²
    """
    
    print(f"\n📈 Создание Excel отчета: {output_path}")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Сводка по элементам"
    
    # Заголовки: материал уже содержит все характеристики (Бетон В30 F150 W6)
    headers = [
        "Тип (RU)", "Тип элемента", "Материал", "Количество, шт", "Объем, м³", "Площадь, м²"
    ]
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Пишем заголовки
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Пишем данные
    row_num = 2
    for idx, item in enumerate(items, 1):
        ws.cell(row=row_num, column=1, value=item['type_ru']).border = thin_border
        ws.cell(row=row_num, column=2, value=item['ifc_class']).border = thin_border
        # Материал уже содержит характеристики: "Бетон В30 F150 W6"
        ws.cell(row=row_num, column=3, value=item['material']).border = thin_border
        ws.cell(row=row_num, column=4, value=item['count']).border = thin_border
        ws.cell(row=row_num, column=5, value=item['volume'] if item['volume'] != '-' else '-').border = thin_border
        ws.cell(row=row_num, column=6, value=item['area'] if item['area'] != '-' else '-').border = thin_border
        row_num += 1
    
    # Автоширина колонок
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)
    
    wb.save(output_path)
    print(f"   ✅ Excel сохранён: {output_path}")
    return output_path


def parse_and_aggregate_specification(session_folder: str) -> Dict[str, Any]:
    """Основная функция: поиск, парсинг, агрегация и сохранение."""
    
    print("\n" + "="*60)
    print("📊 ПАРСИНГ ОТЧЕТА IFC (ifc_report.xlsx)")
    print("="*60)
    
    results = {
        'success': False, 
        'excel_file_found': None, 
        'parsed_items': 0,
        'aggregated_materials': 0, 
        'output_file': None,
        'materials_summary': []
    }
    
    # Шаг 1: Найти файл ifc_report.xlsx
    excel_path = find_ifc_report_file(session_folder)
    if not excel_path:
        print("❌ Файл ifc_report.xlsx не найден")
        return results
    results['excel_file_found'] = excel_path
    
    # Шаг 2: Распарить Excel
    parse_result = parse_excel_report(excel_path)
    if not parse_result['success']:
        results['error'] = parse_result.get('error')
        return results
    results['parsed_items'] = parse_result['total_items']
    results['materials_summary'] = parse_result['items']
    
    # Шаг 3: Создать выходной Excel файл
    output_path = Path(session_folder) / "materials_summary.xlsx"
    create_summary_excel(parse_result['items'], str(output_path))
    results['output_file'] = str(output_path)
    results['success'] = True
    
    # Шаг 4: Сохранить JSON для машинной обработки
    json_path = Path(session_folder) / "materials_summary.json"
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump({
            'success': True,
            'source_file': os.path.basename(excel_path),
            'total_items': parse_result['total_items'],
            'output_excel': 'materials_summary.xlsx',
            'items': parse_result['items']
        }, f, ensure_ascii=False, indent=2)
    
    print(f"\n💾 Данные сохранены в: {json_path}")
    
    # Сводка в консоль
    print("\n" + "="*60)
    print("📊 СВОДКА ПО ЭЛЕМЕНТАМ:")
    print("="*60)
    
    by_type = defaultdict(list)
    for item in parse_result['items']:
        by_type[item['type_ru']].append(item)
    
    for type_ru in sorted(by_type.keys()):
        type_items = by_type[type_ru]
        print(f"\n   [{type_ru}]: {len(type_items)} позиций")
        for it in type_items[:5]:  # Показываем первые 5
            vol_str = f"{it['volume']} м³" if it['volume'] != '-' else '-'
            area_str = f"{it['area']} м²" if it['area'] != '-' else '-'
            print(f"      • {it['ifc_class']} / {it['material']}: {it['count']} шт, V={vol_str}, S={area_str}")
        if len(type_items) > 5:
            print(f"      ... и ещё {len(type_items) - 5}")
    
    # Подсчет количества уникальных видов материалов (тип элемента + материал)
    results['aggregated_materials'] = len(parse_result['items'])
    
    return results


def main(session_folder: str):
    """Основная функция парсинга отчета IFC."""
    print("="*60)
    print("📊 ПАРСИНГ ОТЧЕТА IFC (ifc_report.xlsx)")
    print("="*60)
    print(f"Папка сессии: {session_folder}")
    print("="*60)
    
    results = parse_and_aggregate_specification(session_folder)
    
    return results


if __name__ == "__main__":
    import sys
    
    if len(sys.argv) < 2:
        print("Использование: python xlsx_parser.py <session_folder>")
        print("Пример: python xlsx_parser.py /workspace/uploads/session_id")
        sys.exit(1)
    
    session_folder = sys.argv[1]
    main(session_folder)