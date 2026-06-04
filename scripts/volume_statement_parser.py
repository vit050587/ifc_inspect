#!/usr/bin/env python3
"""
Скрипт парсинга PDF файлов ведомостей объемов (volume_statement).
Извлекает таблицы из PDF файлов и создает сводную таблицу в формате materials_summary.xlsx.

Формат выходной таблицы:
Тип (RU) | Тип элемента | Материал | Количество, шт | Объем, м³ | Площадь, м²

Скрипт работает после xlsx_parser.py и дополняет данные из IFC данными из проектной документации.
"""

import os
import re
import json
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple
from collections import defaultdict
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import fitz  # PyMuPDF


# Mapping русских названий элементов к типам для materials_summary
ELEMENT_TYPE_MAPPING = {
    'шпунт': 'Шпунтовое_ограждение',
    'фундамент': 'Фундаменты',
    'плит': 'Перекрытия',
    'перекрыт': 'Перекрытия',
    'колонн': 'Колонны',
    'стен': 'Стены',
    'лестниц': 'Лестницы',
    'лестнич': 'Лестничные_марши',
    'марш': 'Лестничные_марши',
    'площадк': 'Площадки',
    'вентблок': 'Вентблоки',
    'балок': 'Балки',
    'балка': 'Балки',
    'ригель': 'Балки',
    'ферм': 'Фермы',
    'кровл': 'Крыши',
    'пандус': 'Пандусы',
    'рамп': 'Пандусы',
    'арматур': 'Армирование',
    'сталь': 'Металлоконструкции',
    'металл': 'Металлоконструкции',
    'закладн': 'Закладные_детали',
    'гидроизол': 'Гидроизоляция',
    'гидрошпон': 'Гидрошпонки',
    'пенополистирол': 'Утеплитель',
    'утепл': 'Утеплитель',
    'бетон подготов': 'Бетонная_подготовка',
    'подготов': 'Бетонная_подготовка',
}

# Mapping материалов к стандартным названиям
MATERIAL_NORMALIZATION = {
    'бетон в': 'Бетон',
    'бетон b': 'Бетон',
    'железобетон': 'Железобетон',
    'сборный железобетон': 'Сборный железобетон',
    'сталь': 'Сталь',
    'металл': 'Металлоконструкции',
    'арматура': 'Арматура',
    'пенополистирол': 'Экструдированный пенополистирол',
    'гидроизоляц': 'Гидроизоляция',
    'гидрошпонк': 'Гидрошпонка',
}


def get_element_type_ru(element_name: str) -> str:
    """Определяет русский тип элемента по названию."""
    element_lower = element_name.lower()
    
    for pattern, type_ru in ELEMENT_TYPE_MAPPING.items():
        if pattern in element_lower:
            return type_ru
    
    return 'Прочие_элементы'


def normalize_material_name(material_name: str) -> str:
    """Нормализует название материала."""
    if not material_name:
        return 'Не указан'
    
    material_lower = material_name.lower().strip()
    
    # Проверяем известные паттерны
    for pattern, normalized in MATERIAL_NORMALIZATION.items():
        if pattern in material_lower:
            # Сохраняем характеристики (В30, F150, W6 и т.д.)
            grade_match = re.search(r'([BВ]\d+(?:\.\d+)?)', material_name)
            freeze_match = re.search(r'(F\d+)', material_name)
            water_match = re.search(r'(W\d+)', material_name)
            
            parts = [normalized]
            if grade_match and grade_match.group(1) not in normalized:
                parts.append(grade_match.group(1))
            if freeze_match and freeze_match.group(1) not in normalized:
                parts.append(freeze_match.group(1))
            if water_match and water_match.group(1) not in normalized:
                parts.append(water_match.group(1))
            
            # Добавляем остальные детали (например, "Гравий", "Гранит")
            extra = re.sub(r'[BВ]\d+(?:\.\d+)?|F\d+|W\d+|' + pattern, '', material_lower).strip()
            if extra and len(extra) > 2:
                parts.append(extra.title())
            
            return ' '.join(parts)
    
    return material_name.strip()


def extract_concrete_properties(material_name: str) -> Tuple[Optional[str], Optional[str], Optional[str]]:
    """Извлекает характеристики бетона из названия."""
    if not material_name:
        return None, None, None
    
    grade_match = re.search(r'([BВ]\d+(?:\.\d+)?)', material_name)
    freeze_match = re.search(r'(F\d+)', material_name)
    water_match = re.search(r'(W\d+)', material_name)
    
    grade = grade_match.group(1) if grade_match else None
    freeze = freeze_match.group(1) if freeze_match else None
    water = water_match.group(1) if water_match else None
    
    # Нормализуем класс бетона (латинская B → русская В)
    if grade and grade.startswith('B') and not grade.startswith('В'):
        grade = 'В' + grade[1:]
    
    return grade, freeze, water


def parse_value_with_unit(value_str: str) -> Tuple[float, str]:
    """Парсит значение с единицей измерения (например, '123,45 м³' → (123.45, 'м³'))."""
    if not value_str:
        return 0.0, ''
    
    value_str = str(value_str).strip()
    
    # Паттерн: число (запятая или точка), затем единицы
    match = re.match(r'^([\d,.\s]+)\s*(.*)$', value_str)
    if match:
        num_str = match.group(1).replace(',', '.').replace(' ', '')
        unit = match.group(2).strip()
        try:
            return float(num_str), unit
        except ValueError:
            pass
    
    return 0.0, ''


def extract_tables_from_pdf(pdf_path: str) -> List[Dict[str, Any]]:
    """
    Извлекает таблицы из PDF файла.
    Возвращает список записей с данными о материалах.
    
    Структура таблиц в PDF:
    - Название элемента (например, "Колонна 400х900")
    - Материал (например, "Бетон В30 F150 W8")
    - Объем/Количество (например, "3,40 м³")
    Эти данные могут быть разбиты по разным строкам.
    """
    records = []
    
    try:
        doc = fitz.open(pdf_path)
        print(f"   📄 Обработка PDF: {os.path.basename(pdf_path)} ({len(doc)} стр.)")
        
        for page_num in range(len(doc)):
            page = doc[page_num]
            text = page.get_text()
            
            lines = [line.strip() for line in text.split('\n') if line.strip()]
            
            # Фильтруем служебные строки
            skip_patterns = ['лист', 'изм.', 'кол.', 'подп.', 'дата', 'с-', 'формат', 
                           'согласовано', 'инв.', 'взам.', 'акци', 'обществ', 'генеральн']
            
            clean_lines = []
            for line in lines:
                if any(p in line.lower() for p in skip_patterns):
                    continue
                if line.startswith('С-') and len(line) < 20:  # Шифр документа
                    continue
                clean_lines.append(line)
            
            # Состояние для парсинга табличных данных
            current_element = None
            current_material = None
            
            i = 0
            while i < len(clean_lines):
                line = clean_lines[i]
                
                # Пропускаем заголовки ведомостей
                if 'ведомость' in line.lower() or 'спецификация' in line.lower():
                    current_element = None
                    current_material = None
                    i += 1
                    continue
                
                # Пропускаем заголовки колонок
                if line.lower() in ['марка', 'наименование', 'материал', 'объем', 'площадь', 
                                   'кол.', 'примечание', 'масca', 'число', 'длина, м',
                                   'комплекта', 'чертежей', 'конструкций', 'несущих']:
                    i += 1
                    continue
                
                # Паттерн: элемент с размерами (например, "Колонна 400х900", "ФП H=600")
                elem_dim_pattern = r'^(колonna|фп|балк|ригел|плит|стен|лестниц|марш|площадк|вентблок|гидрошпонк|гидроизол)\s*(.+)?$'
                elem_match = re.match(elem_dim_pattern, line, re.IGNORECASE)
                
                if elem_match and ('х' in line.lower() or 'h=' in line.lower() or '×' in line):
                    current_element = line
                    current_material = None
                    i += 1
                    continue
                
                # Паттерн: материал с характеристиками (Бетон В30 F150 W8, Железобетон, Сталь)
                material_pattern = r'(бетон|[BВ]\d+|F\d+|W\d+|железобетон|сталь|арматура|пенополистирол|гидроизол|гидрошпонк)'
                if re.search(material_pattern, line, re.IGNORECASE):
                    # Проверяем, не является ли строка только объемом
                    vol_only_pattern = r'^[\d,\.\s]+\s*[м³м2мкгтшт]'
                    if not re.match(vol_only_pattern, line):
                        current_material = line
                        i += 1
                        continue
                
                # Паттерн: объем/значение с единицей (619,08 м³, 5370 м², 471,85)
                value_pattern = r'^([\d,\.\s]+)\s*(м³|м2|м²|м|кг|т|шт\.?)?$'
                value_match = re.match(value_pattern, line)
                
                if value_match:
                    value_str = value_match.group(1).strip().replace(',', '.')
                    unit = value_match.group(2) if value_match.group(2) else ''
                    
                    try:
                        value = float(value_str)
                    except ValueError:
                        i += 1
                        continue
                    
                    # Определяем тип значения по единице измерения
                    volume = 0.0
                    area = 0.0
                    weight = 0.0
                    count = 0
                    
                    if unit in ['м³', 'м3']:
                        volume = value
                    elif unit in ['м2', 'м²']:
                        area = value
                    elif unit == 'м':
                        area = value  # Для погонных метров
                    elif unit in ['кг', 'т']:
                        weight = value
                    elif unit in ['шт', 'шт.']:
                        count = int(value)
                    else:
                        # Если нет единицы, предполагаем что это объем
                        volume = value
                    
                    # Создаем запись если есть элемент или материал
                    if current_element or current_material:
                        element_raw = current_element if current_element else ''
                        material_raw = current_material if current_material else element_raw
                        
                        # Если материал не указан, но есть значение - создаем запись с материалом по умолчанию
                        if not current_material and not current_element:
                            i += 1
                            continue
                        
                        material = normalize_material_name(material_raw)
                        grade, freeze, water = extract_concrete_properties(material_raw)
                        
                        # Определяем тип элемента
                        if current_element:
                            element_type = get_element_type_ru(current_element)
                        else:
                            element_type = get_element_type_ru(material_raw)
                        
                        record = {
                            'element_type_raw': element_raw,
                            'type_ru': element_type,
                            'ifc_class': map_element_to_ifc(element_type),
                            'material': material,
                            'concrete_grade': grade,
                            'freeze_durability': freeze,
                            'water_resist': water,
                            'count': count,
                            'volume': volume,
                            'area': area,
                            'weight': weight,
                            'source': os.path.basename(pdf_path),
                            'page': page_num + 1,
                        }
                        
                        records.append(record)
                        
                        # Сбрасываем после использования
                        if volume > 0 or area > 0 or weight > 0 or count > 0:
                            current_element = None
                            current_material = None
                    
                    i += 1
                    continue
                
                # Паттерн: комбинированная строка "Материал Объем" (например, "Бетон В7.5 неармированный 619,08 м³")
                combined_pattern = r'(.+?(?:бетон|железобетон|сталь|арматура|пенополистирол|гидроизол|гидрошпонк).+?)([\d,\.\s]+)\s*(м³|м2|м²|м|кг|т|шт\.?)'
                combined_match = re.search(combined_pattern, line, re.IGNORECASE)
                
                if combined_match:
                    material_raw = combined_match.group(1).strip()
                    value_str = combined_match.group(2).strip().replace(',', '.')
                    unit = combined_match.group(3).strip()
                    
                    try:
                        value = float(value_str)
                    except ValueError:
                        i += 1
                        continue
                    
                    material = normalize_material_name(material_raw)
                    grade, freeze, water = extract_concrete_properties(material_raw)
                    element_type = get_element_type_ru(material_raw)
                    
                    volume = area = weight = count = 0
                    if unit in ['м³', 'м3']:
                        volume = value
                    elif unit in ['м2', 'м²']:
                        area = value
                    elif unit == 'м':
                        area = value
                    elif unit in ['кг', 'т']:
                        weight = value
                    elif unit in ['шт', 'шт.']:
                        count = int(value)
                    
                    record = {
                        'element_type_raw': material_raw,
                        'type_ru': element_type,
                        'ifc_class': map_element_to_ifc(element_type),
                        'material': material,
                        'concrete_grade': grade,
                        'freeze_durability': freeze,
                        'water_resist': water,
                        'count': count,
                        'volume': volume,
                        'area': area,
                        'weight': weight,
                        'source': os.path.basename(pdf_path),
                        'page': page_num + 1,
                    }
                    
                    records.append(record)
                    current_element = None
                    current_material = None
                    i += 1
                    continue
                
                i += 1
        
        doc.close()
        
    except Exception as e:
        print(f"   ❌ Ошибка при чтении PDF {pdf_path}: {e}")
        import traceback
        traceback.print_exc()
    
    return records


def map_element_to_ifc(element_type: str) -> str:
    """Сопоставляет русский тип элемента с IFC классом."""
    mapping = {
        'Стены': 'IfcWall',
        'Колонны': 'IfcColumn',
        'Перекрытия': 'IfcSlab',
        'Фундаменты': 'IfcFooting',
        'Балки': 'IfcBeam',
        'Лестницы': 'IfcStair',
        'Лестничные_марши': 'IfcStairFlight',
        'Пандусы': 'IfcRamp',
        'Крыши': 'IfcRoof',
        'Плиты': 'IfcPlate',
        'Шпунтовое_ограждение': 'IfcPile',
        'Вентблоки': 'IfcBuildingElementProxy',
        'Площадки': 'IfcSlab',
        'Фермы': 'IfcTruss',
        'Армирование': 'IfcReinforcingMesh',
        'Металлоконструкции': 'IfcMember',
        'Закладные_детали': 'IfcFastener',
        'Гидроизоляция': 'IfcMembrane',
        'Гидрошпонки': 'IfcBuildingElementProxy',
        'Утеплитель': 'IfcCovering',
        'Бетонная_подготовка': 'IfcSlab',
        'Прочие_элементы': 'IfcBuildingElementProxy',
    }
    return mapping.get(element_type, 'IfcBuildingElementProxy')


def aggregate_records(records: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Агрегирует записи по ключу (type_ru, ifc_class, material, grade, freeze, water)."""
    aggregated = defaultdict(lambda: {
        'count': 0,
        'volume': 0.0,
        'area': 0.0,
        'weight': 0.0,
        'sources': set(),
    })
    
    for rec in records:
        key = (
            rec['type_ru'],
            rec['ifc_class'],
            rec['material'],
            rec.get('concrete_grade'),
            rec.get('freeze_durability'),
            rec.get('water_resist'),
        )
        
        aggregated[key]['count'] += rec['count']
        aggregated[key]['volume'] += rec['volume']
        aggregated[key]['area'] += rec['area']
        aggregated[key]['weight'] += rec['weight']
        aggregated[key]['sources'].add(rec['source'])
    
    # Преобразуем в список
    items = []
    for (type_ru, ifc_class, material, grade, freeze, water), data in aggregated.items():
        # Формируем полное имя материала с характеристиками
        full_material = material
        if 'бетон' in material.lower() or 'железобетон' in material.lower():
            parts = [material]
            if grade and grade not in material:
                parts.append(grade)
            if freeze and freeze not in material:
                parts.append(freeze)
            if water and water not in material:
                parts.append(water)
            full_material = ' '.join(parts)
        
        item = {
            'type_ru': type_ru,
            'ifc_class': ifc_class,
            'material': full_material,
            'count': int(data['count']) if data['count'] > 0 else '-',
            'volume': round(data['volume'], 3) if data['volume'] > 0 else '-',
            'area': round(data['area'], 3) if data['area'] > 0 else '-',
        }
        items.append(item)
    
    # Сортируем
    items.sort(key=lambda x: (x['type_ru'], x['ifc_class'], x['material']))
    
    return items


def create_summary_excel(items: List[Dict[str, Any]], output_path: str) -> str:
    """Создает Excel файл со сводной таблицей."""
    print(f"\n📈 Создание Excel отчета: {output_path}")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Сводка по материалам"
    
    headers = ["Тип (RU)", "Тип элемента", "Материал", "Количество, шт", "Объем, м³", "Площадь, м²"]
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Заголовки
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Данные
    row_num = 2
    for item in items:
        ws.cell(row=row_num, column=1, value=item['type_ru']).border = thin_border
        ws.cell(row=row_num, column=2, value=item['ifc_class']).border = thin_border
        ws.cell(row=row_num, column=3, value=item['material']).border = thin_border
        ws.cell(row=row_num, column=4, value=item['count']).border = thin_border
        ws.cell(row=row_num, column=5, value=item['volume']).border = thin_border
        ws.cell(row=row_num, column=6, value=item['area']).border = thin_border
        row_num += 1
    
    # Автоширина колонок
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)
    
    wb.save(output_path)
    print(f"   ✅ Excel сохранён: {output_path}")
    return output_path


def find_volume_statement_pdfs(session_folder: str) -> List[str]:
    """Находит PDF файлы в папке volume_statement."""
    volume_folder = Path(session_folder) / 'volume_statement'
    
    if not volume_folder.exists():
        print(f"⚠️  Папка volume_statement не найдена: {volume_folder}")
        return []
    
    pdf_files = list(volume_folder.glob('*.pdf'))
    return [str(f) for f in pdf_files]


def parse_volume_statement_pdfs(session_folder: str) -> Dict[str, Any]:
    """Основная функция: поиск PDF, извлечение таблиц, агрегация и сохранение."""
    
    print("\n" + "="*60)
    print("📊 ПАРСИНГ ВЕДОМОСТЕЙ ОБЪЕМОВ (volume_statement PDF)")
    print("="*60)
    
    results = {
        'success': False,
        'pdf_files_found': [],
        'total_records_extracted': 0,
        'aggregated_items': 0,
        'output_file': None,
        'items': []
    }
    
    # Шаг 1: Найти PDF файлы
    pdf_files = find_volume_statement_pdfs(session_folder)
    if not pdf_files:
        print("❌ PDF файлы в volume_statement не найдены")
        return results
    
    results['pdf_files_found'] = [os.path.basename(f) for f in pdf_files]
    print(f"   📁 Найдено PDF файлов: {len(pdf_files)}")
    
    # Шаг 2: Извлечь данные из всех PDF
    all_records = []
    for pdf_path in pdf_files:
        records = extract_tables_from_pdf(pdf_path)
        all_records.extend(records)
        print(f"   ✓ Извлечено записей: {len(records)}")
    
    results['total_records_extracted'] = len(all_records)
    print(f"\n   📊 Всего извлечено записей: {len(all_records)}")
    
    if not all_records:
        print("⚠️  Таблицы не найдены в PDF файлах")
        return results
    
    # Шаг 3: Агрегировать данные
    aggregated_items = aggregate_records(all_records)
    results['aggregated_items'] = len(aggregated_items)
    results['items'] = aggregated_items
    
    # Шаг 4: Создать Excel файл
    output_path = Path(session_folder) / "volume_statement_summary.xlsx"
    create_summary_excel(aggregated_items, str(output_path))
    results['output_file'] = str(output_path)
    results['success'] = True
    
    # Шаг 5: Сохранить JSON
    json_path = Path(session_folder) / "volume_statement_summary.json"
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump({
            'success': True,
            'source_files': results['pdf_files_found'],
            'total_records': results['total_records_extracted'],
            'aggregated_items': results['aggregated_items'],
            'output_excel': 'volume_statement_summary.xlsx',
            'items': aggregated_items
        }, f, ensure_ascii=False, indent=2)
    
    print(f"\n💾 Данные сохранены в: {json_path}")
    
    # Вывод сводки
    print("\n" + "="*60)
    print("📊 СВОДКА ПО ВЕДОМОСТЯМ ОБЪЕМОВ:")
    print("="*60)
    
    by_type = defaultdict(list)
    for item in aggregated_items:
        by_type[item['type_ru']].append(item)
    
    for type_ru in sorted(by_type.keys()):
        type_items = by_type[type_ru]
        print(f"\n   [{type_ru}]: {len(type_items)} позиций")
        for it in type_items[:5]:
            vol_str = f"{it['volume']} м³" if it['volume'] != '-' else '-'
            area_str = f"{it['area']} м²" if it['area'] != '-' else '-'
            count_str = f"{it['count']} шт" if it['count'] != '-' else '-'
            print(f"      • {it['ifc_class']} / {it['material']}: {count_str}, V={vol_str}, S={area_str}")
        if len(type_items) > 5:
            print(f"      ... и ещё {len(type_items) - 5}")
    
    return results


def main(session_folder: str):
    """Основная функция парсинга ведомостей объемов из PDF."""
    print("="*60)
    print("📊 ПАРСИНГ ВЕДОМОСТЕЙ ОБЪЕМОВ (volume_statement PDF)")
    print("="*60)
    print(f"Папка сессии: {session_folder}")
    print("="*60)
    
    results = parse_volume_statement_pdfs(session_folder)
    return results


if __name__ == "__main__":
    import sys
    
    if len(sys.argv) < 2:
        print("Использование: python volume_statement_parser.py <session_folder>")
        print("Пример: python volume_statement_parser.py /workspace/uploads/session_id")
        sys.exit(1)
    
    session_folder = sys.argv[1]
    main(session_folder)
