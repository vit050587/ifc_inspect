"""
IFC Test Parser - Extracts all constructive elements and their parameters from IFC file.
Generates Excel report with separate sheets for each element type:
  - Columns (Колонны)
  - Slabs (Перекрытия)
  - Walls (Стены)
  - Beams (Балки)
  - Stairs (Лестницы)
  - Ramps (Пандусы)
  - Foundations (Фундаменты) - includes IfcFooting, IfcPile, and IfcSlab with BASESLAB type
  - Other (Остальные)

Each sheet contains ALL parameters for elements of that type.
"""

import ifcopenshell
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime
import os


# ----------------------------------------------------------------------
# Helper functions for extracting property and quantity values
# ----------------------------------------------------------------------

def get_property_value(prop):
    """Extract value from IfcProperty of any type."""
    try:
        if prop.is_a('IfcPropertySingleValue'):
            val = prop.NominalValue
            if val is None:
                return None
            return val.wrappedValue if hasattr(val, 'wrappedValue') else val

        elif prop.is_a('IfcPropertyEnumeratedValue'):
            vals = prop.EnumerationValues
            return ', '.join(str(v) for v in vals) if vals else None

        elif prop.is_a('IfcPropertyListValue'):
            vals = prop.ListValues
            return ', '.join(str(v) for v in vals) if vals else None

        elif prop.is_a('IfcPropertyBoundedValue'):
            lower = prop.LowerBoundValue
            upper = prop.UpperBoundValue
            if lower and upper:
                return f"{lower} – {upper}"
            return lower or upper

        else:   # Other property types - string representation
            if hasattr(prop, 'NominalValue') and prop.NominalValue is not None:
                val = prop.NominalValue
                return val.wrappedValue if hasattr(val, 'wrappedValue') else val
            return str(prop)
    except:
        return None


def get_quantity_value(qty):
    """Extract numeric value from IfcPhysicalQuantity WITHOUT unit conversion."""
    try:
        if qty.is_a('IfcQuantityLength'):
            val = qty.LengthValue
        elif qty.is_a('IfcQuantityArea'):
            val = qty.AreaValue
        elif qty.is_a('IfcQuantityVolume'):
            val = qty.VolumeValue
        elif qty.is_a('IfcQuantityCount'):
            val = qty.CountValue
        elif qty.is_a('IfcQuantityWeight'):
            val = qty.WeightValue
        else:
            val = getattr(qty, 'NominalValue', None)

        if val is None:
            return None
        return val.wrappedValue if hasattr(val, 'wrappedValue') else val
    except:
        return None


def get_element_storey(element, ifc_file):
    """Get the storey/building name that contains this element."""
    try:
        inverses = ifc_file.get_inverse(element)
        for inv in inverses:
            if inv.is_a() == 'IfcRelContainedInSpatialStructure':
                relating = getattr(inv, 'RelatingStructure', None)
                if relating:
                    return getattr(relating, 'Name', None) or getattr(relating, 'LongName', None)
    except:
        pass
    return None


def get_element_properties(element, ifc_file, include_type_props=True):
    """Collect all element parameters (attributes, PropertySet, Quantity, Type).
    Returns dict with keys like 'Element Specific:Long Name', 'PSetName:PropertyName', etc.
    """
    props = {}
    
    # Storey/Building containment
    storey_name = get_element_storey(element, ifc_file)
    if storey_name:
        props['Storey'] = storey_name

    # Basic IfcElement attributes with "Element Specific:" prefix
    base_attrs = ['GlobalId', 'Name', 'Description', 'Tag', 'ObjectType', 'LongName', 'PredefinedType']
    for attr in base_attrs:
        val = getattr(element, attr, None)
        if val is not None:
            props[f'Element Specific:{attr}'] = val
    
    # IFC Class
    props['Ifc Class'] = element.is_a()

    # Properties and quantities directly attached to element (IsDefinedBy)
    if hasattr(element, 'IsDefinedBy'):
        for rel in element.IsDefinedBy or []:
            prop_def = getattr(rel, 'RelatingPropertyDefinition', None)
            if not prop_def:
                continue

            if prop_def.is_a('IfcPropertySet'):
                pset_name = getattr(prop_def, 'Name', 'Unknown')
                for prop in prop_def.HasProperties or []:
                    name = getattr(prop, 'Name', None)
                    if name:
                        key = f'{pset_name}:{name}'
                        if key not in props:
                            props[key] = get_property_value(prop)

            elif prop_def.is_a('IfcElementQuantity'):
                qty_name = getattr(prop_def, 'Name', 'Unknown')
                for qty in prop_def.Quantities or []:
                    name = getattr(qty, 'Name', None)
                    if name:
                        key = f'{qty_name}:{name}'
                        if key not in props:
                            props[key] = get_quantity_value(qty)

    # Properties from element type (IsTypedBy)
    if include_type_props and hasattr(element, 'IsTypedBy'):
        for rel in element.IsTypedBy or []:
            type_obj = getattr(rel, 'RelatingType', None)
            if not type_obj:
                continue

            # Type attributes
            type_attrs = ['GlobalId', 'Name', 'Description', 'ElementType']
            for attr in type_attrs:
                val = getattr(type_obj, attr, None)
                if val is not None and f'Type:{attr}' not in props:
                    props[f'Type:{attr}'] = val

            # PropertySet and Quantity from type
            if hasattr(type_obj, 'HasPropertySets'):
                for ps in type_obj.HasPropertySets or []:
                    if ps.is_a('IfcPropertySet'):
                        pset_name = getattr(ps, 'Name', 'Unknown')
                        for prop in ps.HasProperties or []:
                            name = getattr(prop, 'Name', None)
                            if name:
                                key = f'{pset_name}:{name}'
                                if key not in props:
                                    props[key] = get_property_value(prop)
                    elif ps.is_a('IfcElementQuantity'):
                        qty_name = getattr(ps, 'Name', 'Unknown')
                        for qty in ps.Quantities or []:
                            name = getattr(qty, 'Name', None)
                            if name:
                                key = f'{qty_name}:{name}'
                                if key not in props:
                                    props[key] = get_quantity_value(qty)

    return props


# ----------------------------------------------------------------------
# Element type classification for sheet organization
# ----------------------------------------------------------------------

ELEMENT_TYPE_MAPPING = {
    'IfcColumn': 'Колонны',
    'IfcSlab': 'Перекрытия',
    'IfcWall': 'Стены',
    'IfcBeam': 'Балки',
    'IfcMember': 'Балки',
    'IfcStair': 'Лестницы',
    'IfcStairFlight': 'Лестницы',
    'IfcRamp': 'Пандусы',
    'IfcRampFlight': 'Пандусы',
    'IfcPlate': 'Плиты',
    'IfcFooting': 'Фундаменты',
    'IfcPile': 'Сваи',
    'IfcRoof': 'Кровля',
}

DEFAULT_SHEET_NAME = 'Остальные'


def get_sheet_name_for_element(ifc_class, element=None):
    """Map IFC class to Russian sheet name with special handling for slab types."""
    # Special handling for IfcSlab based on PredefinedType
    if ifc_class == 'IfcSlab' and element:
        predefined_type = getattr(element, 'PredefinedType', None)
        if predefined_type == 'BASESLAB':
            return 'Фундаменты'  # Using simple name without / character
    
    return ELEMENT_TYPE_MAPPING.get(ifc_class, DEFAULT_SHEET_NAME)


# ----------------------------------------------------------------------
# Main parsing and Excel generation function
# ----------------------------------------------------------------------

def parse_ifc_to_sheets(ifc_path, output_excel_path):
    """
    Main entry point: reads IFC and saves Excel report with separate sheets for each element type.
    
    Each sheet contains ALL parameters for elements of that type:
    - Columns: All unique parameter names across all elements of this type
    - Rows: One row per element
    """
    print(f"Opening IFC: {ifc_path}")
    ifc_file = ifcopenshell.open(ifc_path)

    # Collect all structural elements (exclude openings, proxies, etc.)
    structural_types = [
        'IfcColumn', 'IfcSlab', 'IfcWall', 'IfcBeam', 'IfcMember',
        'IfcStair', 'IfcStairFlight', 'IfcRamp', 'IfcRampFlight',
        'IfcPlate', 'IfcFooting', 'IfcPile', 'IfcRoof'
    ]
    
    # Also include building element proxies that might be structural
    all_elements = ifc_file.by_type('IfcElement')
    
    print(f"Total IfcElement count: {len(all_elements)}")

    # Group elements by sheet name
    elements_by_sheet = {}
    
    for i, elem in enumerate(all_elements):
        if i % 500 == 0:
            print(f"Processing element {i} / {len(all_elements)}")
        
        ifc_class = elem.is_a()
        sheet_name = get_sheet_name_for_element(ifc_class, elem)
        
        # Skip non-structural elements for the main sheets
        if sheet_name == DEFAULT_SHEET_NAME and ifc_class not in ['IfcBuildingElementProxy']:
            continue
            
        elem_props = get_element_properties(elem, ifc_file, include_type_props=True)
        
        if sheet_name not in elements_by_sheet:
            elements_by_sheet[sheet_name] = {'elements': [], 'keys': set()}
        
        elements_by_sheet[sheet_name]['elements'].append(elem_props)
        elements_by_sheet[sheet_name]['keys'].update(elem_props.keys())

    print(f"\nElement distribution by sheet:")
    for sheet_name, data in sorted(elements_by_sheet.items()):
        print(f"  {sheet_name}: {len(data['elements'])} elements, {len(data['keys'])} unique parameters")

    # Create Excel workbook
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Create sheet for each element type
    for sheet_name, data in sorted(elements_by_sheet.items()):
        print(f"\nCreating sheet: {sheet_name}")
        ws = wb.create_sheet(sheet_name)
        
        elements = data['elements']
        all_keys = sorted(data['keys'])
        
        # Write headers
        for col_idx, key in enumerate(all_keys, 1):
            cell = ws.cell(row=1, column=col_idx, value=key)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Write element data
        for row_idx, elem_props in enumerate(elements, 2):
            for col_idx, key in enumerate(all_keys, 1):
                val = elem_props.get(key)
                if val is not None:
                    ws.cell(row=row_idx, column=col_idx, value=val).border = thin_border

        # Auto-width columns (max 60 chars)
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    try:
                        max_len = max(max_len, len(str(cell.value)))
                    except:
                        pass
            ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

        print(f"  Written {len(elements)} rows with {len(all_keys)} columns")

    # Save workbook
    wb.save(output_excel_path)
    print(f"\n✅ Excel saved: {output_excel_path}")
    
    return {
        'sheets_created': list(elements_by_sheet.keys()),
        'total_elements': sum(len(d['elements']) for d in elements_by_sheet.values())
    }


# ----------------------------------------------------------------------
# Entry point
# ----------------------------------------------------------------------

if __name__ == "__main__":
    # Configuration
    IFC_FILE = os.path.join(os.path.dirname(__file__), '..', 'data', 'ifc модель КР.ifc')
    OUTPUT_FILE = os.path.join(os.path.dirname(__file__), '..', 'data', 'test.xlsx')
    
    # Normalize paths
    IFC_FILE = os.path.abspath(IFC_FILE)
    OUTPUT_FILE = os.path.abspath(OUTPUT_FILE)
    
    print("=" * 60)
    print("IFC TEST PARSER - Extract all constructive elements")
    print("=" * 60)
    print(f"\nInput IFC:  {IFC_FILE}")
    print(f"Output XLSX: {OUTPUT_FILE}\n")
    
    # Run parser
    result = parse_ifc_to_sheets(IFC_FILE, OUTPUT_FILE)
    
    print("\n" + "=" * 60)
    print("SUMMARY")
    print("=" * 60)
    print(f"Sheets created: {len(result['sheets_created'])}")
    for sheet in result['sheets_created']:
        print(f"  - {sheet}")
    print(f"Total elements processed: {result['total_elements']}")
    print("\n✅ Done!")
