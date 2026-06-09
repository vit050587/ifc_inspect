#!/usr/bin/env python3
"""
IFC Extract - Единый скрипт для извлечения данных из IFC и сохранения всех элементов.
Работает внутри папки сессии.

Очередность операций:
1. extract_ifc_data() - извлекает все данные из IFC файла
2. create_full_elements_excel() - сохраняет полный список элементов в Excel
3. Сохраняет полный список элементов в JSON (elements.json)
4. Извлекает высоту здания и сохраняет в height.txt

Выходные файлы:
- full_elements.xlsx - полный список элементов со всеми параметрами
- elements.json - полный список элементов в формате JSON
- height.txt - высота здания
"""

import ifcopenshell
import ifcopenshell.util.element
import ifcopenshell.util.unit
import json
import os
import math
import re
from collections import defaultdict
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


# Список всех необходимых параметров (колонки)
REQUIRED_COLUMNS = [
    # Element Specific
    "Element Specific:GlobalId",
    "Element Specific:LongName",
    "Element Specific:Name",
    "Element Specific:ObjectType",
    "Element Specific:PredefinedType",
    "Element Specific:Tag",
    # ExpCheckBuilding
    "ExpCheckBuilding:MGE_BuildingAddress",
    "ExpCheckBuilding:MGE_Customer",
    "ExpCheckBuilding:MGE_Designer",
    "ExpCheckBuilding:MGE_ElevationOfRefHeight",
    "ExpCheckBuilding:MGE_ElevationOfTerrain",
    "ExpCheckBuilding:MGE_FunctionalUse",
    "ExpCheckBuilding:MGE_Korpus",
    "ExpCheckBuilding:MGE_NumOfSection",
    "ExpCheckBuilding:MGE_ObjectName",
    "ExpCheckBuilding:MGE_ProjectCode",
    "ExpCheckBuilding:MGE_ProjectName",
    "ExpCheckBuilding:MGE_Section",
    # ExpCheckBuildingStorey
    "ExpCheckBuildingStorey:MGE_ComfortLevel",
    # ExpCheck_Assembly
    "ExpCheck_Assembly:MGE_AssemblyPlace",
    "ExpCheck_Assembly:MGE_Gost",
    "ExpCheck_Assembly:MGE_IsExternal",
    "ExpCheck_Assembly:MGE_Name",
    "ExpCheck_Assembly:MGE_Position",
    # ExpCheck_Beam
    "ExpCheck_Beam:MGE_BeamType",
    "ExpCheck_Beam:MGE_ElementCode",
    "ExpCheck_Beam:MGE_Gost",
    "ExpCheck_Beam:MGE_Material",
    "ExpCheck_Beam:MGE_MaterialCode",
    "ExpCheck_Beam:MGE_Name",
    "ExpCheck_Beam:MGE_Position",
    "ExpCheck_Beam:MGE_SeriaNumber",
    "ExpCheck_Beam:MGE_SteelGrade",
    # ExpCheck_BeamReinforcement
    "ExpCheck_BeamReinforcement:ReinforceStrengthClass",
    # ExpCheck_Column
    "ExpCheck_Column:MGE_ColumnType",
    "ExpCheck_Column:MGE_ElementCode",
    "ExpCheck_Column:MGE_Gost",
    "ExpCheck_Column:MGE_Material",
    "ExpCheck_Column:MGE_MaterialCode",
    "ExpCheck_Column:MGE_Name",
    "ExpCheck_Column:MGE_Position",
    "ExpCheck_Column:MGE_SeriaNumber",
    "ExpCheck_Column:MGE_SteelGrade",
    # ExpCheck_ColumnReinforcement
    "ExpCheck_ColumnReinforcement:MGE_ReinforceStrengthClass",
    # ExpCheck_MaterialConcrete
    "ExpCheck_MaterialConcrete:MGE_ConcreteGost",
    "ExpCheck_MaterialConcrete:MGE_ConcreteGrade",
    "ExpCheck_MaterialConcrete:MGE_FreezeDurability",
    "ExpCheck_MaterialConcrete:MGE_WaterResist",
    # ExpCheck_Ramp
    "ExpCheck_Ramp:MGE_ElementCode",
    "ExpCheck_Ramp:MGE_Gost",
    "ExpCheck_Ramp:MGE_Material",
    "ExpCheck_Ramp:MGE_MaterialCode",
    "ExpCheck_Ramp:MGE_Name",
    "ExpCheck_Ramp:MGE_Position",
    "ExpCheck_Ramp:MGE_Section",
    # ExpCheck_Slab
    "ExpCheck_Slab:MGE_ElementCode",
    "ExpCheck_Slab:MGE_Gost",
    "ExpCheck_Slab:MGE_Material",
    "ExpCheck_Slab:MGE_MaterialCode",
    "ExpCheck_Slab:MGE_Name",
    "ExpCheck_Slab:MGE_Position",
    "ExpCheck_Slab:MGE_SlabType",
    # ExpCheck_SlabReinforcement
    "ExpCheck_SlabReinforcement:MGE_ReinforceStrengthClass",
    # ExpCheck_StairFlight
    "ExpCheck_StairFlight:MGE_ElementCode",
    "ExpCheck_StairFlight:MGE_Gost",
    "ExpCheck_StairFlight:MGE_Material",
    "ExpCheck_StairFlight:MGE_MaterialCode",
    "ExpCheck_StairFlight:MGE_Name",
    "ExpCheck_StairFlight:MGE_Position",
    "ExpCheck_StairFlight:MGE_Section",
    # ExpCheck_Wall
    "ExpCheck_Wall:MGE_ElementCode",
    "ExpCheck_Wall:MGE_Gost",
    "ExpCheck_Wall:MGE_Material",
    "ExpCheck_Wall:MGE_MaterialCode",
    "ExpCheck_Wall:MGE_Name",
    "ExpCheck_Wall:MGE_Position",
    # ExpCheck_WallReinforcement
    "ExpCheck_WallReinforcement:MGE_ReinforceStrengthClass",
    # Ifc Class
    "Ifc Class",
    # Pset_BeamCommon
    "Pset_BeamCommon:IsExternal",
    "Pset_BeamCommon:LoadBearing",
    "Pset_BeamCommon:Reference",
    "Pset_BeamCommon:Roll",
    "Pset_BeamCommon:Slope",
    "Pset_BeamCommon:Span",
    # Pset_BuildingCommon
    "Pset_BuildingCommon:ConstructionMethod",
    "Pset_BuildingCommon:FireProtectionClass",
    "Pset_BuildingCommon:IsLandmarked",
    "Pset_BuildingCommon:NumberOfStoreys",
    "Pset_BuildingCommon:Reference",
    # Pset_BuildingElementProxyCommon
    "Pset_BuildingElementProxyCommon:IsExternal",
    "Pset_BuildingElementProxyCommon:Reference",
    # Pset_BuildingStoreyCommon
    "Pset_BuildingStoreyCommon:AboveGround",
    "Pset_BuildingStoreyCommon:EntranceLevel",
    "Pset_BuildingStoreyCommon:Reference",
    "Pset_BuildingStoreyCommon:SprinklerProtection",
    "Pset_BuildingStoreyCommon:SprinklerProtectionAutomatic",
    # Pset_BuildingSystemCommon
    "Pset_BuildingSystemCommon:Reference",
    # Pset_ColumnCommon
    "Pset_ColumnCommon:IsExternal",
    "Pset_ColumnCommon:LoadBearing",
    "Pset_ColumnCommon:Reference",
    "Pset_ColumnCommon:Slope",
    "Pset_ColumnCommon:ThermalTransmittance",
    # Pset_ConcreteElementGeneral
    "Pset_ConcreteElementGeneral:ConcreteCover",
    "Pset_ConcreteElementGeneral:ConcreteCoverAtLinks",
    "Pset_ConcreteElementGeneral:ConstructionMethod",
    "Pset_ConcreteElementGeneral:ReinforcementVolumeRatio",
    "Pset_ConcreteElementGeneral:StructuralClass",
    # Pset_ElementAssemblyCommon
    "Pset_ElementAssemblyCommon:Reference",
    # Pset_EnvironmentalImpactIndicators
    "Pset_EnvironmentalImpactIndicators:Reference",
    # Pset_ManufacturerTypeInformation
    "Pset_ManufacturerTypeInformation:AssemblyPlace",
    "Pset_ManufacturerTypeInformation:Manufacturer",
    # Pset_MemberCommon
    "Pset_MemberCommon:IsExternal",
    "Pset_MemberCommon:LoadBearing",
    "Pset_MemberCommon:Reference",
    # Pset_OpeningElementCommon
    "Pset_OpeningElementCommon:Reference",
    # Pset_PlateCommon
    "Pset_PlateCommon:IsExternal",
    "Pset_PlateCommon:LoadBearing",
    "Pset_PlateCommon:Reference",
    # Pset_RampCommon
    "Pset_RampCommon:IsExternal",
    "Pset_RampCommon:Reference",
    # Pset_RampFlightCommon
    "Pset_RampFlightCommon:Reference",
    "Pset_RampFlightCommon:Slope",
    # Pset_ReinforcementBarPitchOfBeam
    "Pset_ReinforcementBarPitchOfBeam:Reference",
    # Pset_ReinforcementBarPitchOfColumn
    "Pset_ReinforcementBarPitchOfColumn:Reference",
    # Pset_ReinforcementBarPitchOfSlab
    "Pset_ReinforcementBarPitchOfSlab:Reference",
    # Pset_ReinforcementBarPitchOfWall
    "Pset_ReinforcementBarPitchOfWall:Reference",
    # Pset_SlabCommon
    "Pset_SlabCommon:IsExternal",
    "Pset_SlabCommon:LoadBearing",
    "Pset_SlabCommon:PitchAngle",
    "Pset_SlabCommon:Reference",
    "Pset_SlabCommon:ThermalTransmittance",
    # Pset_StairFlightCommon
    "Pset_StairFlightCommon:IsExternal",
    "Pset_StairFlightCommon:LoadBearing",
    "Pset_StairFlightCommon:Reference",
    # Pset_WallCommon
    "Pset_WallCommon:ExtendToStructure",
    "Pset_WallCommon:IsExternal",
    "Pset_WallCommon:LoadBearing",
    "Pset_WallCommon:Reference",
    "Pset_WallCommon:ThermalTransmittance",
    # Qto_BeamBaseQuantities
    "Qto_BeamBaseQuantities:CrossSectionArea",
    "Qto_BeamBaseQuantities:GrossSurfaceArea",
    "Qto_BeamBaseQuantities:GrossVolume",
    "Qto_BeamBaseQuantities:Length",
    "Qto_BeamBaseQuantities:NetSurfaceArea",
    "Qto_BeamBaseQuantities:NetVolume",
    "Qto_BeamBaseQuantities:OuterSurfaceArea",
    # Qto_ColumnBaseQuantities
    "Qto_ColumnBaseQuantities:CrossSectionArea",
    "Qto_ColumnBaseQuantities:GrossVolume",
    "Qto_ColumnBaseQuantities:Length",
    "Qto_ColumnBaseQuantities:NetVolume",
    "Qto_ColumnBaseQuantities:OuterSurfaceArea",
    # Qto_MemberBaseQuantities
    "Qto_MemberBaseQuantities:CrossSectionArea",
    "Qto_MemberBaseQuantities:GrossSurfaceArea",
    "Qto_MemberBaseQuantities:GrossVolume",
    "Qto_MemberBaseQuantities:Length",
    "Qto_MemberBaseQuantities:NetSurfaceArea",
    "Qto_MemberBaseQuantities:NetVolume",
    "Qto_MemberBaseQuantities:OuterSurfaceArea",
    # Qto_OpeningElementBaseQuantities
    "Qto_OpeningElementBaseQuantities:Area",
    "Qto_OpeningElementBaseQuantities:Depth",
    "Qto_OpeningElementBaseQuantities:Height",
    "Qto_OpeningElementBaseQuantities:Width",
    # Qto_SlabBaseQuantities
    "Qto_SlabBaseQuantities:GrossArea",
    "Qto_SlabBaseQuantities:GrossVolume",
    "Qto_SlabBaseQuantities:NetArea",
    "Qto_SlabBaseQuantities:NetVolume",
    "Qto_SlabBaseQuantities:Perimeter",
    "Qto_SlabBaseQuantities:Width",
    # Qto_StairFlightBaseQuantities
    "Qto_StairFlightBaseQuantities:GrossVolume",
    "Qto_StairFlightBaseQuantities:Length",
    "Qto_StairFlightBaseQuantities:NetVolume",
    # Qto_WallBaseQuantities
    "Qto_WallBaseQuantities:GrossSideArea",
    "Qto_WallBaseQuantities:GrossVolume",
    "Qto_WallBaseQuantities:Height",
    "Qto_WallBaseQuantities:Length",
    "Qto_WallBaseQuantities:NetSideArea",
    "Qto_WallBaseQuantities:NetVolume",
    "Qto_WallBaseQuantities:Width",
    # Storey
    "Storey",
    # Type
    "Type:GlobalId",
    "Type:Name",
]


def get_unit_scale(ifc_file):
    """Получить масштаб единиц из IFC файла"""
    try:
        units = ifcopenshell.util.unit.get_unit_scale(ifc_file)
        return units
    except:
        return 1.0


def calculate_volume_by_geometry(shape, unit_scale):
    """Расчет объема через bounding box с учетом масштаба"""
    try:
        bbox_min = shape.bounding_box.lower_corner
        bbox_max = shape.bounding_box.upper_corner
        
        dx = (bbox_max[0] - bbox_min[0]) * unit_scale
        dy = (bbox_max[1] - bbox_min[1]) * unit_scale
        dz = (bbox_max[2] - bbox_min[2]) * unit_scale
        
        return abs(dx * dy * dz)
    except Exception as e:
        return 0.0


def calculate_area_by_geometry(shape, unit_scale):
    """Расчет площади поверхности через геометрию"""
    try:
        area = shape.geometry.surface_area
        return area * (unit_scale ** 2)
    except Exception as e:
        return 0.0


def get_property_value(psets, pset_name, prop_name):
    """Получить значение свойства из psets"""
    if pset_name in psets:
        props = psets[pset_name]
        if prop_name in props:
            val = props[prop_name]
            if val is not None and prop_name != 'id':
                return val
    return None


def extract_element_properties(element, psets):
    """Извлечь все свойства элемента - только те, что есть в модели"""
    data = {}
    
    # Element Specific properties (базовые атрибуты элемента)
    if getattr(element, 'GlobalId', None):
        data["Element Specific:GlobalId"] = getattr(element, 'GlobalId', None)
    if getattr(element, 'LongName', None):
        data["Element Specific:LongName"] = getattr(element, 'LongName', None)
    if getattr(element, 'Name', None):
        data["Element Specific:Name"] = getattr(element, 'Name', None)
    if getattr(element, 'ObjectType', None):
        data["Element Specific:ObjectType"] = getattr(element, 'ObjectType', None)
    if getattr(element, 'PredefinedType', None):
        data["Element Specific:PredefinedType"] = getattr(element, 'PredefinedType', None)
    if getattr(element, 'Tag', None):
        data["Element Specific:Tag"] = getattr(element, 'Tag', None)
    
    # Динамическое извлечение всех параметров из psets
    for pset_name, props in psets.items():
        for prop_name, prop_value in props.items():
            if prop_value is not None and prop_name != 'id':
                # Формируем ключ в формате "PsetName:PropertyName"
                key = f"{pset_name}:{prop_name}"
                data[key] = prop_value
    
    # Ifc Class
    data["Ifc Class"] = element.is_a()
    
    return data


def classify_element(element, name, psets=None):
    """Классификация элемента на более детальные подтипы"""
    element_type = element.is_a()
    name_upper = name.upper() if name else ""
    
    # Получаем информацию о бетоне для проверки класса
    concrete_class = ""
    if psets:
        full_text = f"{name} "
        for pset_name, pset_data in psets.items():
            for prop_name, prop_val in pset_data.items():
                if isinstance(prop_val, str):
                    full_text += f"{prop_val} "
        class_match = re.search(r'[BВ]\s?(\d+(?:[.,]\d+)?)', full_text.upper())
        if class_match:
            concrete_class = class_match.group(1).replace(',', '.')
    
    if element_type == "IfcSlab":
        # Приоритет 1: Лестничные площадки (должны быть отдельно)
        # Проверяем по имени элемента и по типу
        if ("ЛЕСТНИЧ" in name_upper or "STAIR" in name_upper or 
            "ПЛОЩАДК" in name_upper or "ПЛОЩ." in name_upper or
            "ЛМ" in name_upper):
            return "Лестница_Площадка"
        
        # Приоритет 2: Подготовка фундамента (низкие классы бетона, аномальные классы или явные маркеры)
        # Классы бетона ниже В20 или аномальные (например В230) считаем подготовкой
        try:
            class_value = float(concrete_class) if concrete_class else 0
            # Считаем подготовкой: класс < 20 ИЛИ класс > 100 (аномалия типа В230)
            if class_value > 0 and (class_value < 20 or class_value > 100):
                return "Фундамент_Подготовка_Бетон"
        except ValueError:
            pass
        
        if "ПОДГОТОВКА" in name_upper or "СТЯЖКА" in name_upper:
            if "ПЕСОК" in name_upper or "SAND" in name_upper or "ЦЕМЕНТНО-ПЕСЧАНАЯ" in name_upper:
                return "Фундамент_Подготовка_Песок"
            elif "ЩЕБЕНЬ" in name_upper or "GRAVEL" in name_upper or "CRUSHED" in name_upper:
                return "Фундамент_Подготовка_Щебень"
            else:
                return "Фундамент_Подготовка_Бетон"
        
        # Приоритет 3: Фундаментные плиты (ищем явные маркеры фундамента)
        if ("ФУНДАМЕНТ" in name_upper and "ПОДГОТОВКА" not in name_upper) or \
           "FOUNDATION SLAB" in name_upper or \
           "BASE SLAB" in name_upper or \
           "ФУНДАМЕНТНАЯ ПЛИТА" in name_upper:
            return "Фундамент_Плита"
        
        if "ФП" in name_upper and "ПЕРЕКРЫТИЕ" not in name_upper and "ПОДГОТОВКА" not in name_upper:
            return "Фундамент_Плита"
        
        # Приоритет 4: Перекрытия (по умолчанию для плит)
        return "Перекрытие_Плита"
    
    if element_type == "IfcWall":
        return "Стена"
    if element_type == "IfcColumn":
        return "Колонна"
    if element_type in ["IfcStair", "IfcStairFlight"]:
        return "Лестница_Марш"
    if element_type == "IfcBeam":
        return "Балка"
        
    return element_type


def parse_concrete_properties(name, psets):
    """Парсинг свойств бетона из имени и PSets"""
    props = {
        "material": "Бетон",
        "class": "",
        "frost": "",
        "water": "",
        "reinforced": True,
        "type_detail": ""
    }
    
    full_text = f"{name} "
    for pset_name, pset_data in psets.items():
        for prop_name, prop_val in pset_data.items():
            if isinstance(prop_val, str):
                full_text += f"{prop_val} "
    
    full_text_upper = full_text.upper()
    
    # Проверка на пенополистирол/утеплитель
    if "ПЕНОПОЛИСТИРОЛ" in full_text_upper or "ЭКСТРУДИРОВАНН" in full_text_upper or "УТЕПЛИТЕЛ" in full_text_upper:
        props["material"] = "Экструдированный пенополистирол"
        props["type_detail"] = "Экструдированный пенополистирол"
        props["reinforced"] = False
        return props
    
    # Проверка на не бетонные материалы подготовки
    if "ПЕСОК" in full_text_upper or "SAND" in full_text_upper:
        props["material"] = "Песок"
        props["type_detail"] = "Песок"
        props["reinforced"] = False
    elif "ЩЕБЕНЬ" in full_text_upper or "GRAVEL" in full_text_upper or "CRUSHED STONE" in full_text_upper:
        props["material"] = "Щебень"
        props["type_detail"] = "Щебень"
        props["reinforced"] = False
    elif "ПОДГОТОВКА" in full_text_upper and ("В7.5" in full_text_upper or "B7.5" in full_text_upper or "НЕАРМИРОВАННЫЙ" in full_text_upper):
        props["type_detail"] = "Бетон неармированный (подготовка)"
        props["reinforced"] = False

    # Поиск класса бетона
    class_match = re.search(r'[BВ]\s?(\d+(?:[.,]\d+)?)', full_text_upper)
    if class_match:
        val = class_match.group(1).replace(',', '.')
        props["class"] = f"В{val}"
        if float(val) < 15:
             props["reinforced"] = False

    # Поиск морозостойкости
    frost_match = re.search(r'F\s?(\d+)', full_text_upper)
    if frost_match:
        props["frost"] = f"F{frost_match.group(1)}"
    
    # Поиск водонепроницаемости
    water_match = re.search(r'W\s?(\d+)', full_text_upper)
    if water_match:
        props["water"] = f"W{water_match.group(1)}"
        
    return props


def extract_ifc_data(ifc_path):
    """Основная функция извлечения данных из IFC"""
    ifc_file = ifcopenshell.open(ifc_path)
    unit_scale = get_unit_scale(ifc_file)

    elements_data = []
    all_elements = ifc_file.by_type("IfcProduct")

    print(f"Всего элементов в файле: {len(all_elements)}")

    for i, element in enumerate(all_elements):
        if not element.Representation:
            continue

        name = element.Name if element.Name else "Без имени"
        tag = element.Tag if element.Tag else ""
        element_type = element.is_a()

        psets = {}
        try:
            pset_definitions = ifcopenshell.util.element.get_psets(element)
            for pset_name, props in pset_definitions.items():
                psets[pset_name] = props
        except:
            pass

        # Извлекаем все параметры из модели
        item = extract_element_properties(element, psets)

        # Добавляем вычисляемые значения объема и площади
        volume = 0.0
        area = 0.0
        shape = None

        # Приоритет №1: Объем из свойств
        found_volume_in_props = False
        for pset_name, props in psets.items():
            if "NetVolume" in props:
                vol = props["NetVolume"]
                if isinstance(vol, (int, float)):
                    volume = vol
                    found_volume_in_props = True
                    break
            if "GrossVolume" in props and not found_volume_in_props:
                vol = props["GrossVolume"]
                if isinstance(vol, (int, float)):
                    if vol > 10000:
                        volume = vol / 1000.0
                    else:
                        volume = vol
                    found_volume_in_props = True
                    break

        # Если объема нет в свойствах, пробуем рассчитать по геометрии
        if not found_volume_in_props:
            try:
                settings = ifcopenshell.geom.settings()
                settings.set(settings.USE_PYTHON_OPENCASCADE, True)
                shape = ifcopenshell.geom.create_shape(settings, element)
                if shape:
                    volume = calculate_volume_by_geometry(shape, unit_scale)
            except Exception as e:
                pass

        # Расчет площади
        try:
            if shape:
                area = calculate_area_by_geometry(shape, unit_scale)
            else:
                settings = ifcopenshell.geom.settings()
                settings.set(settings.USE_PYTHON_OPENCASCADE, True)
                shape = ifcopenshell.geom.create_shape(settings, element)
                if shape:
                    area = calculate_area_by_geometry(shape, unit_scale)
        except:
            pass

        # Добавляем volume и area в item
        item["volume_m3"] = round(volume, 4)
        item["area_m2"] = round(area, 4)

        # Классификация и материал
        category = classify_element(element, name, psets)
        material_props = parse_concrete_properties(name, psets)

        item["category"] = category
        item["material"] = material_props["material"]
        item["concrete_class"] = material_props["class"]
        item["frost_resistance"] = material_props["frost"]
        item["water_permeability"] = material_props["water"]
        item["is_reinforced"] = material_props["reinforced"]
        item["material_detail"] = material_props["type_detail"]

        elements_data.append(item)

        if (i + 1) % 500 == 0:
            print(f"Обработано {i + 1} элементов...")

    return elements_data


def get_russian_type_name(category, ifc_type):
    """Получение русского названия типа элемента"""
    type_mapping = {
        "Балка": "Балки",
        "Колонна": "Колонны",
        "Стена": "Стены",
        "Перекрытие_Плита": "Перекрытия",
        "Фундамент_Плита": "Фундаментные_плиты",
        "Фундамент_Подготовка_Бетон": "Подготовка_фундамента",
        "Фундамент_Подготовка_Песок": "Подготовка_фундамента",
        "Фундамент_Подготовка_Щебень": "Подготовка_фундамента",
        "Лестница_Марш": "Лестничные_марши",
        "Лестница_Площадка": "Лестничные_площадки",
        "Пандус": "Пандусы",
    }

    # Сначала пробуем по категории
    if category in type_mapping:
        return type_mapping[category]

    # Затем по IFC типу
    ifc_type_mapping = {
        "IfcBeam": "Балки",
        "IfcColumn": "Колонны",
        "IfcWall": "Стены",
        "IfcSlab": "Перекрытия",
        "IfcFooting": "Фундаментные_плиты",
        "IfcStairFlight": "Лестничные_марши",
        "IfcStair": "Лестницы",
        "IfcRamp": "Пандусы",
        "IfcPlate": "Плиты",
        "IfcMember": "Элементы_каркаса",
        "IfcFurnishingElement": "Мебель",
        "IfcBuildingElementProxy": "Прочие_элементы",
        "IfcOpeningElement": "Проемы",
        "IfcCurtainWall": "Навесные_стены",
        "IfcWindow": "Окна",
        "IfcDoor": "Двери",
        "IfcRoof": "Крыши",
        "IfcPile": "Сваи",
        "IfcRailing": "Ограждения",
        "IfcCovering": "Покрытия",
        "IfcReinforcingMesh": "Арматурные_сетки",
        "IfcBuildingStorey": "Этажи",
        "IfcBuilding": "Здание",
    }

    if ifc_type in ifc_type_mapping:
        return ifc_type_mapping[ifc_type]

    return "Прочие"


def format_material_string(material_props):
    """Форматирование строки материала с учетом класса и свойств"""
    material = material_props.get("material", "")
    concrete_class = material_props.get("class", "")
    frost = material_props.get("frost", "")
    water = material_props.get("water", "")
    type_detail = material_props.get("type_detail", "")
    
    # Если есть type_detail, используем его
    if type_detail:
        return type_detail
    
    # Собираем полное описание материала
    parts = [material] if material else []
    
    if concrete_class:
        parts.append(concrete_class)
    
    if frost:
        parts.append(frost)
    
    if water:
        parts.append(water)
    
    # Если материал пустой, но есть свойства бетона
    if not parts and (concrete_class or frost or water):
        props = []
        if concrete_class:
            props.append(concrete_class)
        if frost:
            props.append(frost)
        if water:
            props.append(water)
        return " ".join(props)
    
    return " ".join(parts) if parts else "-"


def create_summary_table(data):
    """Создание сводной таблицы в требуемом формате"""
    df = pd.DataFrame(data)

    # Список IFC классов для исключения (неинформативные элементы)
    excluded_ifc_classes = [
        "IfcDiscreteAccessory",
        "IfcElementAssembly",
        "IfcOpeningElement",
        "IfcBuildingElementProxy"
    ]

    # Фильтрация по IFC классам
    df = df[~df["Ifc Class"].isin(excluded_ifc_classes)]

    # Добавляем колонку с русским названием типа
    df["Тип (RU)"] = df.apply(
        lambda row: get_russian_type_name(row.get("category", ""), row.get("Ifc Class", "")),
        axis=1
    )

    # Добавляем колонку с полным описанием материала
    df["Материал_полный"] = df.apply(
        lambda row: format_material_string({
            "material": row.get("material", ""),
            "class": row.get("concrete_class", ""),
            "frost": row.get("frost_resistance", ""),
            "water": row.get("water_permeability", ""),
            "type_detail": row.get("material_detail", "")
        }),
        axis=1
    )

    # Фильтрация строк с неинформативными материалами
    def is_informative_material(row):
        material_full = row["Материал_полный"]
        concrete_class = row.get("concrete_class", "")
        material_detail = row.get("material_detail", "")
        ifc_class = row.get("Ifc Class", "")

        # ИСКЛЮЧЕНИЕ: Лестницы и лестничные марши всегда включаем в таблицу
        if ifc_class in ["IfcStair", "IfcStairFlight"]:
            return True

        # Если есть detail-описание - это информативно
        if material_detail:
            return True

        # Если есть класс бетона (В25, В30, В7.5 и т.п.) - это информативно
        if concrete_class and re.search(r'В\d+\.?\d*', concrete_class):
            return True

        # Если материал содержит конкретное название (не просто "Бетон")
        if material_full and material_full != "-" and len(material_full.split()) > 1:
            return True

        return False

    df = df[df.apply(is_informative_material, axis=1)]

    # Группировка по типу, IFC классу и материалу
    group_cols = ["Тип (RU)", "Ifc Class", "Материал_полный"]

    summary = df.groupby(group_cols, dropna=False).agg(
        count=("id" if "id" in df.columns else "Element Specific:GlobalId", "count"),
        total_volume_m3=("volume_m3", "sum"),
        total_area_m2=("area_m2", "sum")
    ).reset_index()

    # Переименовываем колонки в соответствии с требуемым форматом
    summary = summary.rename(columns={
        "Ifc Class": "Тип элемента",
        "Материал_полный": "Материал",
        "count": "Количество, шт",
        "total_volume_m3": "Объем, м³"
    })

    # Форматирование числовых значений
    summary["Объем, м³"] = summary["Объем, м³"].apply(lambda x: round(x, 3) if pd.notna(x) and x != 0 else None)

    # Замена None на "-" для отображения
    summary["Объем, м³"] = summary["Объем, м³"].fillna("-")

    # Сортировка
    summary = summary.sort_values(by=["Тип (RU)", "Тип элемента", "Материал"], ascending=[True, True, True])

    # Выбираем только нужные колонки в правильном порядке (без Площади)
    summary = summary[["Тип (RU)", "Тип элемента", "Материал", "Количество, шт", "Объем, м³"]]

    return summary


def create_summary_excel(items, output_path):
    """Создание Excel файла со сводной таблицей"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Материалы"
    
    # Заголовки
    headers = ["Тип (RU)", "Тип элемента", "Материал", "Количество, шт", "Объем, м³"]
    
    # Стили
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Записываем заголовки
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Записываем данные
    for row_idx, item in enumerate(items, 2):
        ws.cell(row=row_idx, column=1, value=item.get("Тип (RU)", "")).border = thin_border
        ws.cell(row=row_idx, column=2, value=item.get("Тип элемента", "")).border = thin_border
        ws.cell(row=row_idx, column=3, value=item.get("Материал", "")).border = thin_border
        ws.cell(row=row_idx, column=4, value=item.get("Количество, шт", 0)).border = thin_border
        
        vol = item.get("Объем, м³", "-")
        ws.cell(row=row_idx, column=5, value=vol).border = thin_border
    
    # Авто-ширина колонок
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                try:
                    max_len = max(max_len, len(str(cell.value)))
                except:
                    pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)
    
    wb.save(output_path)
    print(f"✅ Сводная таблица сохранена: {output_path}")


def create_full_elements_excel(data, output_path):
    """Создание Excel файла с полным списком элементов и их параметрами"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Все элементы"
    
    # Основные колонки согласно требованиям
    main_headers = [
        "Наименование",
        "Код IFC",
        "Подземный/Надземный",
        "Ширина, м",
        "Высота, м",
        "Длина, м",
        "Толщина, м",
        "Объем, м³",
        "Площадь, м²",
        "Материал",
        "Характеристики материала"
    ]
    
    # Добавляем все остальные колонки из данных элементов
    # Собираем все уникальные ключи из всех элементов
    all_data_keys = set()
    for item in data:
        all_data_keys.update(item.keys())
    
    # Формируем заголовки: основные + все остальные из данных
    all_headers = main_headers + [col for col in sorted(all_data_keys) if col not in main_headers]
    
    # Стили
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Записываем заголовки
    for col_idx, header in enumerate(all_headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Записываем данные
    for row_idx, item in enumerate(data, 2):
        # Наименование
        name = item.get("Element Specific:Name") or item.get("Element Specific:LongName") or ""
        ws.cell(row=row_idx, column=1, value=name).border = thin_border
        
        # Код IFC
        ifc_code = item.get("ExpCheck_Beam:MGE_ElementCode") or \
                   item.get("ExpCheck_Column:MGE_ElementCode") or \
                   item.get("ExpCheck_Slab:MGE_ElementCode") or \
                   item.get("ExpCheck_Wall:MGE_ElementCode") or \
                   item.get("ExpCheck_Ramp:MGE_ElementCode") or \
                   item.get("ExpCheck_StairFlight:MGE_ElementCode") or ""
        ws.cell(row=row_idx, column=2, value=ifc_code).border = thin_border
        
        # Подземный/Надземный
        above_ground = item.get("Pset_BuildingStoreyCommon:AboveGround")
        storey_elevation = None
        # Пытаемся определить по этажу
        if hasattr(item.get("Storey"), 'Elevation'):
            storey_elevation = item["Storey"].Elevation
        elif isinstance(item.get("Storey"), (int, float)):
            storey_elevation = item["Storey"]
        
        if storey_elevation is not None:
            underground_status = "Подземный" if storey_elevation < 0 else "Надземный"
        elif above_ground is not None:
            underground_status = "Надземный" if above_ground else "Подземный"
        else:
            underground_status = "Не определено"
        ws.cell(row=row_idx, column=3, value=underground_status).border = thin_border
        
        # Размеры (извлекаем из Qto quantities)
        width = item.get("Qto_OpeningElementBaseQuantities:Width") or \
                item.get("Qto_SlabBaseQuantities:Width") or \
                item.get("Qto_WallBaseQuantities:Width") or ""
        height = item.get("Qto_OpeningElementBaseQuantities:Height") or \
                 item.get("Qto_WallBaseQuantities:Height") or ""
        length = item.get("Qto_BeamBaseQuantities:Length") or \
                 item.get("Qto_ColumnBaseQuantities:Length") or \
                 item.get("Qto_MemberBaseQuantities:Length") or \
                 item.get("Qto_SlabBaseQuantities:Width") or \
                 item.get("Qto_WallBaseQuantities:Length") or ""
        thickness = item.get("Qto_WallBaseQuantities:Width") or \
                    item.get("Qto_SlabBaseQuantities:Width") or ""
        
        ws.cell(row=row_idx, column=4, value=width).border = thin_border
        ws.cell(row=row_idx, column=5, value=height).border = thin_border
        ws.cell(row=row_idx, column=6, value=length).border = thin_border
        ws.cell(row=row_idx, column=7, value=thickness).border = thin_border
        
        # Объем и площадь
        ws.cell(row=row_idx, column=8, value=item.get("volume_m3", "")).border = thin_border
        ws.cell(row=row_idx, column=9, value=item.get("area_m2", "")).border = thin_border
        
        # Материал
        material = item.get("material", "")
        concrete_class = item.get("concrete_class", "")
        frost = item.get("frost_resistance", "")
        water = item.get("water_permeability", "")
        
        material_parts = [material] if material else []
        if concrete_class:
            material_parts.append(concrete_class)
        if frost:
            material_parts.append(frost)
        if water:
            material_parts.append(water)
        
        material_str = " ".join(material_parts) if material_parts else ""
        ws.cell(row=row_idx, column=10, value=material_str).border = thin_border
        
        # Характеристики материала (полное описание)
        material_detail = item.get("material_detail", "")
        char_parts = []
        if concrete_class:
            char_parts.append(f"Класс бетона: {concrete_class}")
        if frost:
            char_parts.append(f"Морозостойкость: {frost}")
        if water:
            char_parts.append(f"Водонепроницаемость: {water}")
        if material_detail:
            char_parts.append(f"Тип: {material_detail}")
        
        char_str = "; ".join(char_parts) if char_parts else ""
        ws.cell(row=row_idx, column=11, value=char_str).border = thin_border
        
        # Остальные колонки из данных элементов
        for col_idx, header in enumerate(all_headers[11:], 12):
            value = item.get(header, "")
            ws.cell(row=row_idx, column=col_idx + 11, value=value).border = thin_border
    
    # Авто-ширина колонок
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                try:
                    max_len = max(max_len, len(str(cell.value)))
                except:
                    pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)
    
    wb.save(output_path)
    print(f"✅ Полный список элементов сохранен: {output_path}")


def extract_building_info(ifc_file):
    """
    Извлечь информацию об объекте строительства из открытого IFC файла.
    
    Args:
        ifc_file: Открытый IFC файл
        
    Returns:
        Словарь с информацией о здании
    """
    result = {
        "project": {},
        "building": {},
        "storeys": [],
        "summary": {}
    }
    
    # === Извлечение данных проекта ===
    projects = ifc_file.by_type("IfcProject")
    if projects:
        project = projects[0]
        result["project"]["name"] = getattr(project, 'Name', None)
        result["project"]["long_name"] = getattr(project, 'LongName', None)
        result["project"]["description"] = getattr(project, 'Description', None)
    
    # === Извлечение данных здания ===
    buildings = ifc_file.by_type("IfcBuilding")
    if buildings:
        building = buildings[0]
        
        # Базовые атрибуты
        result["building"]["name"] = getattr(building, 'Name', None)
        result["building"]["long_name"] = getattr(building, 'LongName', None)
        result["building"]["elevation_ref_height"] = getattr(building, 'ElevationOfRefHeight', None)
        result["building"]["elevation_terrain"] = getattr(building, 'ElevationOfTerrain', None)
        
        # Извлечение свойств через IsDefinedBy
        psets_data = {}
        if hasattr(building, 'IsDefinedBy'):
            for rel in building.IsDefinedBy:
                if hasattr(rel, 'RelatingPropertyDefinition'):
                    pset_def = rel.RelatingPropertyDefinition
                    pset_name = getattr(pset_def, 'Name', 'Unknown')
                    
                    if hasattr(pset_def, 'HasProperties'):
                        for prop in pset_def.HasProperties:
                            prop_name = getattr(prop, 'Name', None)
                            prop_value = getattr(prop, 'NominalValue', None)
                            
                            if prop_name and prop_value is not None:
                                # Извлекаем само значение из NominalValue
                                if hasattr(prop_value, 'wrappedValue'):
                                    actual_value = prop_value.wrappedValue
                                else:
                                    actual_value = prop_value
                                
                                if pset_name not in psets_data:
                                    psets_data[pset_name] = {}
                                psets_data[pset_name][prop_name] = actual_value
        
        # Сохраняем все найденные свойства
        result["building"]["properties"] = psets_data
        
        # Ключевые параметры для summary
        if "ExpCheck_Building" in psets_data:
            exp_props = psets_data["ExpCheck_Building"]
            result["building"]["address"] = exp_props.get("MGE_BuildingAddress")
            result["building"]["customer"] = exp_props.get("MGE_Customer")
            result["building"]["designer"] = exp_props.get("MGE_Designer")
            result["building"]["project_name"] = exp_props.get("MGE_ProjectName")
            result["building"]["object_name"] = exp_props.get("MGE_ObjectName")
            result["building"]["project_code"] = exp_props.get("MGE_ProjectCode")
            result["building"]["korpus"] = exp_props.get("MGE_Korpus")
            result["building"]["section"] = exp_props.get("MGE_Section")
            result["building"]["num_of_section"] = exp_props.get("MGE_NumOfSection")
            result["building"]["functional_use"] = exp_props.get("MGE_FunctionalUse")
            
            # Высота от нулевой отметки (разница между референсной высотой и уровнем земли)
            ref_height = exp_props.get("MGE_ElevationOfRefHeight")
            terrain_height = exp_props.get("MGE_ElevationOfTerrain")
            if ref_height is not None and terrain_height is not None:
                # Обычно MGE_ElevationOfRefHeight - это абсолютная высота в мм
                # Конвертируем в метры
                result["building"]["ref_height_m"] = float(ref_height) / 1000.0 if ref_height else None
                result["building"]["terrain_height_m"] = float(terrain_height) / 1000.0 if terrain_height else None
        
        if "Pset_BuildingCommon" in psets_data:
            common_props = psets_data["Pset_BuildingCommon"]
            result["building"]["number_of_storeys"] = common_props.get("NumberOfStoreys")
            result["building"]["construction_method"] = common_props.get("ConstructionMethod")
            result["building"]["fire_protection_class"] = common_props.get("FireProtectionClass")
            result["building"]["is_landmarked"] = common_props.get("IsLandmarked")
    
    # === Извлечение информации об этажах ===
    storeys = ifc_file.by_type("IfcBuildingStorey")
    above_ground_storeys = []
    below_ground_storeys = []
    
    for storey in storeys:
        elevation = getattr(storey, 'Elevation', None)
        storey_info = {
            "name": getattr(storey, 'Name', None),
            "elevation": float(elevation) / 1000.0 if elevation is not None else None  # Конвертация в метры
        }
        
        # Определяем, надземный или подземный этаж
        if elevation is not None:
            if elevation > 0:
                above_ground_storeys.append(storey_info)
            else:
                below_ground_storeys.append(storey_info)
        else:
            above_ground_storeys.append(storey_info)
        
        result["storeys"].append(storey_info)
    
    # Сортировка этажей по высоте
    result["storeys"].sort(key=lambda x: x["elevation"] if x["elevation"] is not None else 0)
    
    # === Формирование сводной информации (summary) ===
    total_storeys = len(above_ground_storeys) + len(below_ground_storeys)
    
    # Расчет высоты здания
    max_elevation = None
    min_elevation = None
    
    for storey in result["storeys"]:
        elev = storey["elevation"]
        if elev is not None:
            if max_elevation is None or elev > max_elevation:
                max_elevation = elev
            if min_elevation is None or elev < min_elevation:
                min_elevation = elev
    
    # Общая высота здания от нулевой отметки
    if max_elevation is not None:
        result["summary"]["total_height_from_zero_m"] = round(max_elevation, 3)
    
    # Глубина подземной части
    if min_elevation is not None and min_elevation < 0:
        result["summary"]["underground_depth_m"] = round(abs(min_elevation), 3)
    
    # Полная высота (от низа подземной части до верха)
    if max_elevation is not None and min_elevation is not None:
        result["summary"]["total_height_full_m"] = round(max_elevation - min_elevation, 3)
    
    # Этажность
    result["summary"]["above_ground_storeys"] = len(above_ground_storeys)
    result["summary"]["below_ground_storeys"] = len(below_ground_storeys)
    result["summary"]["total_storeys"] = total_storeys
    
    # Если есть данные из Pset_BuildingCommon
    if result["building"].get("number_of_storeys"):
        result["summary"]["storeys_from_pset"] = result["building"]["number_of_storeys"]
    
    # Адрес
    result["summary"]["address"] = result["building"].get("address")
    
    return result


def parse_ifc_file(ifc_path, output_folder):
    """
    Основная функция для сервиса ifc_inspect.
    Работает внутри папки сессии.
    
    Очередность:
    1. Извлекаем данные из IFC (extract_ifc_data)
    2. Сохраняем полный список элементов в JSON и Excel
    3. Извлекаем высоту здания и сохраняем в height.txt
    
    Args:
        ifc_path: Path to IFC file
        output_folder: Folder to save results (session folder)
        
    Returns:
        Dict with parsing results
    """
    print(f"\n{'='*60}")
    print("📊 ОБРАБОТКА IFC МОДЕЛИ")
    print(f"{'='*60}")
    print(f"IFC файл: {os.path.basename(ifc_path)}")
    print(f"Папка сессии: {output_folder}")
    
    # Шаг 1: Извлечение данных из IFC
    print("\n📋 Шаг 1: Извлечение данных из IFC...")
    data = extract_ifc_data(ifc_path)
    print(f"   ✅ Извлечено {len(data)} элементов")
    
    # Шаг 1.5: Извлечение информации о здании
    ifc_file = ifcopenshell.open(ifc_path)
    building_info = extract_building_info(ifc_file)
    
    # Шаг 2: Сохранение полного списка элементов в Excel
    print("\n📋 Шаг 2: Сохранение полного списка элементов...")
    full_elements_path = Path(output_folder) / "full_elements.xlsx"
    create_full_elements_excel(data, str(full_elements_path))
    
    # Шаг 3: Сохранение полного списка элементов в JSON
    print("\n📋 Шаг 3: Сохранение полного списка элементов в JSON...")
    json_path = Path(output_folder) / "elements.json"
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump({
            'success': True,
            'source_file': os.path.basename(ifc_path),
            'total_elements': len(data),
            'elements': data
        }, f, ensure_ascii=False, indent=2, default=str)
    print(f"   ✅ Сохранено {len(data)} элементов в elements.json")
    
    # Шаг 4: Сохранение высоты здания
    print("\n📏 Шаг 4: Сохранение высоты здания...")
    height_m = building_info.get("summary", {}).get("total_height_from_zero_m")
    if height_m is not None:
        height_path = Path(output_folder) / "height.txt"
        with open(height_path, 'w', encoding='utf-8') as f:
            f.write(str(height_m))
        print(f"   ✅ Высота здания: {height_m} м (сохранено в height.txt)")
    else:
        print("   ⚠️ Не удалось определить высоту здания")
        height_m = None
    
    result = {
        'success': True,
        'total_elements': len(data),
        'height_m': height_m,
        'full_elements_file': 'full_elements.xlsx',
        'elements_json_file': 'elements.json',
        'height_file': 'height.txt'
    }
    
    print(f"\n💾 Результаты сохранены в: {output_folder}")
    print(f"   • full_elements.xlsx - полный список элементов")
    print(f"   • elements.json - полный список элементов (JSON)")
    print(f"   • height.txt - высота здания")
    print(f"\n✅ Обработка IFC завершена. Всего элементов: {len(data)}")
    
    return result


# Entry point when run directly
if __name__ == "__main__":
    import sys
    
    if len(sys.argv) < 2:
        print("Использование: python ifc_parser.py <path_to_IFC> [output_folder]")
        print("Пример: python ifc_parser.py /workspace/data/model.ifc /workspace/uploads/session_id")
        sys.exit(1)
    
    ifc_file = sys.argv[1]
    output_folder = sys.argv[2] if len(sys.argv) > 2 else "."
    
    parse_ifc_file(ifc_file, output_folder)
