#!/usr/bin/env python3
"""
Скрипт для фильтрации Перечня работ КР по высоте здания.

Логика работы:
1. Читает число из файла height.txt в папке сессии
2. Открывает файл data/Перечень работ КР.xlsx, лист "ВОР КР+расценки"
3. В столбце F находит строки с переменной "H" (например: "H<30 м", "30<H<40 м", "t>300 мм\n57<H<75 м")
4. Подставляет высоту из height.txt вместо "H" и проверяет условие
5. Оставляет только строки, где условие выполняется ИЛИ где нет условий с H
6. Сохраняет новый файл Перечень работ КР_new.xlsx в папке сессии
"""

import os
import re
import shutil
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


def parse_height_condition(condition_text, building_height):
    """
    Парсит условие с переменной H и проверяет его для заданной высоты.
    
    Args:
        condition_text: Строка условия, например "H<30 м", "30<H<40 м", "t>300 мм\n57<H<75 м"
        building_height: Числовая высота здания в метрах
    
    Returns:
        True если условие выполняется, False иначе
    """
    if not condition_text:
        return True  # Если нет условия, считаем что подходит
    
    # Проверяем, есть ли вообще H в тексте
    if 'H' not in condition_text:
        return True  # Нет H - значит условие не связано с высотой, пропускаем
    
    # Разбиваем на строки (может быть несколько условий через \n)
    lines = condition_text.split('\n')
    
    for line in lines:
        line = line.strip()
        if not line:
            continue
        
        # Ищем паттерны с H - порядок важен! Сначала сложные, потом простые
        
        # Паттерн 3: число<H<число (например "30<H<40 м") - должен быть раньше чем H<число
        match = re.search(r'(\d+(?:\.\d+)?)\s*<\s*H\s*<\s*(\d+(?:\.\d+)?)', line)
        if match:
            min_height = float(match.group(1))
            max_height = float(match.group(2))
            if not (min_height < building_height < max_height):
                return False
            continue
        
        # Паттерн 4: число<=H<=число или смешанные варианты
        match = re.search(r'(\d+(?:\.\d+)?)\s*<=?\s*H\s*<=?\s*(\d+(?:\.\d+)?)', line)
        if match:
            min_height = float(match.group(1))
            max_height = float(match.group(2))
            if not (min_height <= building_height <= max_height):
                return False
            continue
        
        # Паттерн 5: H>=число
        match = re.search(r'H\s*>=\s*(\d+(?:\.\d+)?)', line)
        if match:
            min_height = float(match.group(1))
            if building_height < min_height:
                return False
            continue
        
        # Паттерн 6: H<=число
        match = re.search(r'H\s*<=\s*(\d+(?:\.\d+)?)', line)
        if match:
            max_height = float(match.group(1))
            if building_height > max_height:
                return False
            continue
        
        # Паттерн 1: H<число (например "H<30 м")
        match = re.search(r'H\s*<\s*(\d+(?:\.\d+)?)', line)
        if match:
            max_height = float(match.group(1))
            if building_height >= max_height:
                return False
            continue
        
        # Паттерн 2: H>число (например "H>30 м")
        match = re.search(r'H\s*>\s*(\d+(?:\.\d+)?)', line)
        if match:
            min_height = float(match.group(1))
            if building_height <= min_height:
                return False
            continue
    
    return True


def filter_works_by_height(session_folder, source_excel_path=None):
    """
    Фильтрует перечень работ по высоте здания.
    
    Args:
        session_folder: Папка сессии
        source_excel_path: Путь к исходному файлу (по умолчанию data/Перечень работ КР.xlsx)
    
    Returns:
        Dict с результатами
    """
    print(f"\n{'='*60}")
    print("📐 ФИЛЬТРАЦИЯ ПЕРЕЧНЯ РАБОТ ПО ВЫСОТЕ ЗДАНИЯ")
    print(f"{'='*60}")
    
    # Шаг 1: Чтение высоты из height.txt
    height_file = Path(session_folder) / "height.txt"
    if not height_file.exists():
        print(f"❌ Файл {height_file} не найден")
        return {'success': False, 'error': 'height.txt not found'}
    
    with open(height_file, 'r', encoding='utf-8') as f:
        try:
            building_height = float(f.read().strip())
        except ValueError as e:
            print(f"❌ Ошибка чтения высоты: {e}")
            return {'success': False, 'error': f'Invalid height value: {e}'}
    
    print(f"📏 Высота здания: {building_height} м")
    
    # Шаг 2: Определение пути к исходному файлу
    if source_excel_path is None:
        source_excel_path = Path(__file__).parent.parent / "data" / "Перечень работ КР.xlsx"
    
    if not Path(source_excel_path).exists():
        print(f"❌ Файл {source_excel_path} не найден")
        return {'success': False, 'error': 'Source Excel file not found'}
    
    print(f"📁 Исходный файл: {source_excel_path}")
    
    # Шаг 3: Загрузка workbook
    wb = load_workbook(str(source_excel_path), data_only=True)
    
    # Работаем только с листом "ВОР КР+расценки"
    if "ВОР КР+расценки" not in wb.sheetnames:
        print(f"❌ Лист 'ВОР КР+расценки' не найден. Доступные листы: {wb.sheetnames}")
        return {'success': False, 'error': 'Sheet not found'}
    
    ws_source = wb["ВОР КР+расценки"]
    
    # Шаг 4: Сначала собираем все строки в память для предварительной обработки
    all_rows = list(ws_source.iter_rows(values_only=True))
    
    # Шаг 5: Предварительная обработка - заполнение значений столбцов A, B, C, D
    # Алгоритм: 
    # 1. Находим заголовки подразделов (строки начинающиеся с "Подраздел" или "Раздел")
    # 2. Извлекаем номер подраздела из заголовка (например "1" из "Подраздел 1. ...")
    # 3. Для первой строки с кодом расценки после заголовка сохраняем значения B, C, D
    # 4. Все последующие строки до следующего заголовка получают:
    #    - A = номер подраздела
    #    - B, C, D = значения из первой строки работы этого подраздела
    processed_rows = []
    current_subsection_number = None  # Текущий номер подраздела (например "1", "2", "3.1" и т.д.)
    current_b_value = None  # Значение столбца B для текущего подраздела
    current_c_value = None  # Значение столбца C для текущего подраздела
    current_d_value = None  # Значение столбца D для текущего подраздела
    first_work_row_in_subsection = True  # Флаг: первая ли это строка работы в подразделе
    
    for row_idx, row in enumerate(all_rows):
        row_list = list(row) if row else []
        
        # Пропускаем заголовок (первую строку)
        if row_idx == 0:
            processed_rows.append(row_list)
            continue
        
        col_a = row_list[0] if len(row_list) > 0 else None
        col_g = row_list[6] if len(row_list) > 6 else None
        
        # Проверяем, является ли строка заголовком раздела или подраздела
        if col_a and isinstance(col_a, str) and (col_a.startswith('Подраздел') or col_a.startswith('Раздел')):
            # Извлекаем номер подраздела из заголовка
            # Например: "Подраздел 1. Надземная часть здания. Стены" -> "1"
            # Или: "Раздел 5. Монолитные ж/б конструкции" -> "5"
            match = re.match(r'(?:Подраздел|Раздел)\s+(\d+(?:\.\d+)?)', col_a)
            if match:
                current_subsection_number = match.group(1)
            else:
                current_subsection_number = None
            # Сбрасываем значения B, C, D для нового подраздела
            current_b_value = None
            current_c_value = None
            current_d_value = None
            first_work_row_in_subsection = True
            processed_rows.append(row_list)
            continue
        
        # Если есть код расценки в столбце G - это строка работы
        if col_g is not None and col_g != '':
            new_row = row_list[:]
            
            if current_subsection_number is not None:
                # Устанавливаем номер подраздела в столбец A
                new_row[0] = current_subsection_number
                
                # Если это первая строка работы в подразделе, запоминаем B, C, D
                if first_work_row_in_subsection:
                    current_b_value = new_row[1] if len(new_row) > 1 else None
                    current_c_value = new_row[2] if len(new_row) > 2 else None
                    current_d_value = new_row[3] if len(new_row) > 3 else None
                    first_work_row_in_subsection = False
                else:
                    # Для всех последующих строк устанавливаем B, C, D как в первой строке
                    if len(new_row) > 1:
                        new_row[1] = current_b_value
                    if len(new_row) > 2:
                        new_row[2] = current_c_value
                    if len(new_row) > 3:
                        new_row[3] = current_d_value
            
            processed_rows.append(new_row)
        else:
            # Строки без кода расценки (промежуточные заголовки, пустые строки)
            # Сохраняем как есть, не сбрасывая current_subsection_number
            processed_rows.append(row_list)
    
    # Шаг 6: Фильтрация строк по высоте
    print("\n🔍 Фильтрация строк...")
    
    # Создаем новый workbook для результата
    wb_result = Workbook()
    
    # Копируем все листы из исходного файла
    for sheet_name in wb.sheetnames:
        if sheet_name == "ВОР КР+расценки":
            # Этот лист будем фильтровать
            ws_result = wb_result.active
            ws_result.title = "ВОР КР+расценки"
        else:
            # Просто копируем остальные листы
            ws_source_other = wb[sheet_name]
            ws_result_other = wb_result.create_sheet(title=sheet_name)
            for row in ws_source_other.iter_rows(values_only=True):
                ws_result_other.append(list(row) if row else [])
            continue
        
        # Копируем стили и данные из исходного листа с фильтрацией
        filtered_row_count = 0
        total_row_count = 0
        
        for row_idx, row in enumerate(processed_rows, 1):
            total_row_count += 1
            
            # Пропускаем заголовок (первую строку)
            if row_idx == 1:
                ws_result.append(row)
                filtered_row_count += 1
                continue
            
            # Проверяем, есть ли в строке столбец F (индекс 5)
            if len(row) <= 5:
                # Если столбца F нет, просто копируем строку
                ws_result.append(row)
                filtered_row_count += 1
                continue
            
            f_value = row[5]  # Столбец F (0-индексированный)
            
            # Если в столбце F есть текст с H, проверяем условие
            if f_value and isinstance(f_value, str) and 'H' in f_value:
                if parse_height_condition(f_value, building_height):
                    ws_result.append(row)
                    filtered_row_count += 1
                # else: строка отфильтрована
            else:
                # Если нет H в столбце F, копируем строку
                ws_result.append(row)
                filtered_row_count += 1
        
        print(f"   Обработано строк: {total_row_count - 1}")
        print(f"   Осталось после фильтрации: {filtered_row_count - 1}")
    
    # Шаг 7: Сохранение результата
    output_path = Path(session_folder) / "Перечень работ КР_new.xlsx"
    
    # Применяем стили (копируем ширины колонок из оригинала)
    ws_source = wb["ВОР КР+расценки"]
    for col_letter in ws_source.column_dimensions:
        ws_result.column_dimensions[col_letter].width = ws_source.column_dimensions[col_letter].width
    
    wb_result.save(str(output_path))
    print(f"\n✅ Результат сохранен: {output_path}")
    
    return {
        'success': True,
        'output_file': 'Перечень работ КР_new.xlsx',
        'output_path': str(output_path),
        'building_height': building_height,
        'source_file': str(source_excel_path)
    }


# Entry point when run directly
if __name__ == "__main__":
    import sys
    
    if len(sys.argv) < 2:
        print("Использование: python filter_works_by_height.py <session_folder> [source_excel_path]")
        print("Пример: python filter_works_by_height.py /workspace/uploads/session_id")
        sys.exit(1)
    
    session_folder = sys.argv[1]
    source_path = sys.argv[2] if len(sys.argv) > 2 else None
    
    result = filter_works_by_height(session_folder, source_path)
    
    if result.get('success'):
        print(f"\n✅ Фильтрация завершена успешно!")
        print(f"   Высота здания: {result['building_height']} м")
        print(f"   Результат: {result['output_file']}")
    else:
        print(f"\n❌ Ошибка: {result.get('error', 'Unknown error')}")
        sys.exit(1)
